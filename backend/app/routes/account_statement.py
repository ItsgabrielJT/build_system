from __future__ import annotations

from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import Response

from app.auth.dependencies import get_current_user
from app.config.database import get_db
from app.repositories.delinquency_repository import DelinquencyRepository
from app.repositories.owner_repository import OwnerRepository
from app.services.account_statement_service import AccountStatementService

router = APIRouter(tags=["account-statement"])


@router.get("/account-statement")
async def get_account_statement(
    start_period: Optional[str] = None,
    end_period: Optional[str] = None,
    owner_id: Optional[UUID] = None,
    user: dict = Depends(get_current_user),
    db=Depends(get_db),
):
    service = AccountStatementService(DelinquencyRepository(db), OwnerRepository(db))
    resolved_id = await service.resolve_owner_id(user, owner_id)
    if not resolved_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Propietario no encontrado para este usuario",
        )
    return await service.get_statement(resolved_id, start_period, end_period)


@router.get("/account-statement/export")
async def export_account_statement(
    format: str = "csv",
    start_period: Optional[str] = None,
    end_period: Optional[str] = None,
    owner_id: Optional[UUID] = None,
    user: dict = Depends(get_current_user),
    db=Depends(get_db),
):
    service = AccountStatementService(DelinquencyRepository(db), OwnerRepository(db))
    resolved_id = await service.resolve_owner_id(user, owner_id)
    if not resolved_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Propietario no encontrado para este usuario",
        )

    rows = await service.get_statement(resolved_id, start_period, end_period)

    headers_row = ["Período", "Departamento", "Esperado", "Multas", "Pagado", "Saldo", "Estado"]

    if format == "excel":
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estado de Cuenta"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="123C7A", end_color="123C7A", fill_type="solid")
        for col_idx, header in enumerate(headers_row, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            ws.append([
                row["period"],
                row["apartment_code"],
                row["esperado"],
                row["multas"],
                row["pagado"],
                row["saldo"],
                row["status"],
            ])

        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = "estado-cuenta.xlsx"

    elif format == "pdf":
        content = await service.statement_pdf(resolved_id, start_period, end_period)
        media_type = "application/pdf"
        filename = "estado-cuenta.pdf"

    else:
        import csv as csv_module
        import io

        output = io.StringIO()
        writer = csv_module.writer(output)
        writer.writerow(headers_row)
        for row in rows:
            writer.writerow([
                row["period"],
                row["apartment_code"],
                row["esperado"],
                row["multas"],
                row["pagado"],
                row["saldo"],
                row["status"],
            ])
        content = output.getvalue().encode("utf-8-sig")
        media_type = "text/csv"
        filename = "estado-cuenta.csv"

    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/account-statement/expense-certificate")
async def export_expense_certificate(
    owner_id: Optional[UUID] = None,
    user: dict = Depends(get_current_user),
    db=Depends(get_db),
):
    service = AccountStatementService(DelinquencyRepository(db), OwnerRepository(db))
    resolved_id = await service.resolve_owner_id(user, owner_id)
    if not resolved_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Propietario no encontrado para este usuario",
        )
    try:
        content = await service.expense_certificate_pdf(resolved_id)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc)) from exc
    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="certificado-expensas.pdf"'},
    )
