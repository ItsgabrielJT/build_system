from __future__ import annotations

import calendar
import csv
import io
import math
import re
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Optional
from uuid import UUID
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm, inch
from reportlab.platypus import HRFlowable, KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.graphics.shapes import Drawing, Rect, String, Line, Circle

from app.repositories.expense_repository import ExpenseRepository
from app.repositories.income_repository import IncomeRepository
from app.repositories.payment_repository import PaymentRepository
from app.services.pdf_branding import (
    build_pdf_brand_header,
    build_pdf_footer_bar,
    build_pdf_signature_seal_qr_grid,
    get_building_contact_lines,
    get_building_logo,
    get_building_name,
    get_default_building_config,
)
from app.services.delinquency_service import DelinquencyService

_MONTH_PERIOD_RE = re.compile(r"^\d{4}-\d{2}$")
_CONFIRMED_PAYMENT_STATUSES = {"REGISTRADO", "APROBADO"}
_EXCLUDED_EXPENSE_STATUSES = {"ANULADO", "ANULADA"}
_PDF_BLUE = colors.HexColor("#123c7a")
_PDF_NAVY = colors.HexColor("#00316f")
_PDF_BORDER = colors.HexColor("#b9cbe3")
_PDF_LIGHT = colors.HexColor("#f8fafc")
_PDF_PALE_BLUE = colors.HexColor("#eaf3ff")


class ReportService:
    def __init__(
        self,
        delinquency_service: DelinquencyService,
        payment_repo: PaymentRepository,
        expense_repo: ExpenseRepository,
        income_repo: Optional[IncomeRepository] = None,
    ) -> None:
        self._delinquency = delinquency_service
        self._payment_repo = payment_repo
        self._expense_repo = expense_repo
        self._income_repo = income_repo

    def _validate_month_period(self, period: str) -> str:
        if not _MONTH_PERIOD_RE.match(period):
            raise ValueError("Período debe tener formato YYYY-MM")
        return period

    def _previous_period(self, period: str) -> str:
        month_start = datetime.strptime(f"{period}-01", "%Y-%m-%d").date()
        previous_month_last_day = month_start - timedelta(days=1)
        return previous_month_last_day.strftime("%Y-%m")

    def _is_confirmed_payment(self, payment: dict) -> bool:
        return (payment.get("status") or "").upper() in _CONFIRMED_PAYMENT_STATUSES

    def _is_valid_expense(self, expense: dict) -> bool:
        return (expense.get("status") or "REGISTRADO").upper() not in _EXCLUDED_EXPENSE_STATUSES

    def _is_valid_income(self, income: dict) -> bool:
        return (income.get("status") or "REGISTRADO").upper() == "REGISTRADO"

    async def _monthly_payments(self, period: str) -> list[dict]:
        period = self._validate_month_period(period)
        historical = await self._payment_repo.get_all(payment_month=period, status="REGISTRADO")
        approved = await self._payment_repo.get_all(payment_month=period, status="APROBADO")
        payments_by_id: dict[str, dict] = {}
        for payment in [*historical, *approved]:
            payment_id = str(payment.get("id") or "")
            payments_by_id[payment_id] = payment
        return [
            payment
            for payment in payments_by_id.values()
            if self._is_confirmed_payment(payment)
        ]

    async def _monthly_expenses(self, period: str) -> list[dict]:
        period = self._validate_month_period(period)
        expenses_payload = await self._expense_repo.get_by_month(period)
        expenses = expenses_payload.get("data", []) if isinstance(expenses_payload, dict) else expenses_payload
        return [expense for expense in expenses if self._is_valid_expense(expense)]

    async def _monthly_incomes(self, period: str) -> list[dict]:
        if not self._income_repo:
            return []
        period = self._validate_month_period(period)
        incomes = await self._income_repo.get_all(period=period, status="REGISTRADO")
        return [income for income in incomes if self._is_valid_income(income)]

    def _sum_amount(self, rows: list[dict]) -> Decimal:
        return sum(Decimal(str(row.get("amount", 0) or 0)) for row in rows)

    def _normalize_payment_income(self, payment: dict) -> dict:
        return {
            **payment,
            "income_source_type": "payment",
            "income_label": "Pago registrado",
            "income_concept": f"Pago {payment.get('period') or ''}".strip(),
            "income_date": payment.get("paid_at"),
            "income_category": "Alícuotas y pagos",
        }

    def _normalize_manual_income(self, income: dict) -> dict:
        return {
            **income,
            "paid_at": income.get("date"),
            "income_source_type": "income",
            "income_label": income.get("source") or income.get("category") or "Ingreso manual",
            "income_concept": income.get("concept") or "Ingreso",
            "income_date": income.get("date"),
            "income_category": income.get("category") or "Otros ingresos",
            "owner_name": income.get("owner_name") or "Ingreso general",
            "apartment_code": income.get("apartment_code") or "",
            "period": income.get("period") or (
                income.get("date").strftime("%Y-%m") if hasattr(income.get("date"), "strftime") else ""
            ),
            "method": income.get("method") or income.get("source") or "",
            "status": income.get("status") or "REGISTRADO",
        }

    async def _income_entries(
        self,
        period: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> list[dict]:
        payments = [self._normalize_payment_income(payment) for payment in await self._payments(period, start_date, end_date)]
        incomes = [self._normalize_manual_income(income) for income in await self._incomes(period, start_date, end_date)]
        return [*payments, *incomes]

    def _build_breakdown(self, rows: list[dict], key: str, fallback: str) -> list[dict]:
        totals: dict[str, Decimal] = {}
        for row in rows:
            label = row.get(key) or fallback
            totals[label] = totals.get(label, Decimal("0")) + Decimal(str(row.get("amount", 0) or 0))
        return [
            {"label": label, "amount": amount}
            for label, amount in sorted(totals.items(), key=lambda item: item[1], reverse=True)
        ]

    def _variation_percent(self, current: Decimal, previous: Decimal) -> Optional[float]:
        if previous <= 0:
            return None
        return round(float((current - previous) / previous * 100), 2)

    def _money(self, value) -> str:
        return f"${float(Decimal(str(value or 0))):,.2f}"

    def _usd(self, value) -> str:
        return f"USD {float(Decimal(str(value or 0))):,.2f}"

    def _period_name(self, period: Optional[str]) -> str:
        if not period:
            return "Todos"
        try:
            parsed = datetime.strptime(f"{period}-01", "%Y-%m-%d")
        except ValueError:
            return period
        months = [
            "enero", "febrero", "marzo", "abril", "mayo", "junio",
            "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
        ]
        return f"{months[parsed.month - 1]} {parsed.year}"

    def _spanish_date(self, value: date) -> str:
        months = [
            "enero", "febrero", "marzo", "abril", "mayo", "junio",
            "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
        ]
        return f"{value.day:02d} de {months[value.month - 1]} de {value.year}"

    def _delinquency_status_label(self, status: str | None) -> str:
        labels = {
            "OVERDUE": "Vencido",
            "CURRENT": "Al día",
        }
        return labels.get((status or "").upper(), status or "")

    def _p(self, text: str, size: int = 8, *, bold: bool = False, color="#102a56", align: str = "LEFT", raw: bool = False) -> Paragraph:
        safe_text = str(text or "") if raw else escape(str(text or ""))
        return Paragraph(
            f'<font color="{color}">{"<b>" if bold else ""}{safe_text}{"</b>" if bold else ""}</font>',
            ParagraphStyle(
                f"PdfP{size}{bold}{align}",
                fontName="Helvetica-Bold" if bold else "Helvetica",
                fontSize=size,
                leading=size + 2,
                alignment={"LEFT": 0, "CENTER": 1, "RIGHT": 2}.get(align, 1),
            ),
        )

    def _pdf_table_text(self, text: str, *, chunk_size: int = 16) -> str:
        """Add break opportunities to long table values before ReportLab lays them out."""
        normalized = re.sub(r"\s+", " ", str(text or "").strip())
        if not normalized:
            return ""

        def split_token(token: str) -> str:
            if len(token) <= chunk_size:
                return token
            parts = re.split(r"([_./\\-])", token)
            if len(parts) > 1:
                token = "".join(
                    part + (" " if part in {"_", ".", "/", "\\", "-"} else "")
                    for part in parts
                )
            return " ".join(token[i:i + chunk_size] for i in range(0, len(token), chunk_size))

        return " ".join(split_token(token) for token in normalized.split(" "))

    def _table_p(
        self,
        text: str,
        size: int = 7,
        *,
        bold: bool = False,
        color="#102a56",
        align: str = "LEFT",
    ) -> Paragraph:
        safe_text = escape(self._pdf_table_text(text))
        return Paragraph(
            f'<font color="{color}">{"<b>" if bold else ""}{safe_text}{"</b>" if bold else ""}</font>',
            ParagraphStyle(
                f"PdfTableP{size}{bold}{align}",
                fontName="Helvetica-Bold" if bold else "Helvetica",
                fontSize=size,
                leading=size + 2,
                alignment={"LEFT": 0, "CENTER": 1, "RIGHT": 2}.get(align, 0),
                splitLongWords=1,
                wordWrap="CJK",
            ),
        )

    async def _modern_report_header(self, title: str, subtitle: str, width: float, *, building: Optional[dict] = None) -> list:
        building = building if building is not None else await get_default_building_config(getattr(self._payment_repo, "_conn", None))
        generated = datetime.now()
        logo = get_building_logo(building, max_width=2.8 * cm, max_height=2.4 * cm)
        if not logo:
            logo = self._p(
                "EDIFICIO<br/><font size='16'><b>TORRES</b></font><br/>NETANYA",
                10,
                bold=True,
                color="#092b62",
                raw=True,
            )
        info_width = 4.1 * cm
        info = Table(
            [
                [self._p("Formatos: Habittauio", 8, bold=True, align="LEFT")],
                [self._p("Fecha de impresión", 8, bold=True, align="LEFT")],
                [self._p(self._spanish_date(generated.date()), 8, align="LEFT")],
            ],
            colWidths=[info_width],
        )
        info.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.8, _PDF_BLUE),
            ("LINEBELOW", (0, 0), (-1, 0), 0.6, _PDF_BORDER),
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        title_block = self._p(
            f"<font size='22'><b>{escape(title)}</b></font><br/><font size='8'>{escape(subtitle)}</font>",
            12,
            color="#082f6f",
            align="LEFT",
            raw=True,
        )
        header = Table([[logo, title_block, info]], colWidths=[3.3 * cm, width - 7.4 * cm, info_width])
        header.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 1.2, _PDF_BLUE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING", (2, 0), (2, 0), 0),
            ("RIGHTPADDING", (2, 0), (2, 0), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ("LINEAFTER", (0, 0), (0, 0), 0.8, _PDF_BLUE),
            ("LINEAFTER", (1, 0), (1, 0), 0.8, _PDF_BLUE),
        ]))
        return [header, Spacer(1, 0.25 * cm)]

    async def _three_column_report_header(
        self,
        title: str,
        subtitle: str,
        width: float,
        *,
        building: Optional[dict] = None,
        right_text: Optional[str] = None,
    ) -> list:
        building = building if building is not None else await get_default_building_config(getattr(self._payment_repo, "_conn", None))
        logo = get_building_logo(building, max_width=2.8 * cm, max_height=2.4 * cm)
        if not logo:
            logo = self._p(
                f"<b>{escape(get_building_name(building).upper())}</b>",
                10,
                bold=True,
                align="CENTER",
                raw=True,
            )

        title_html = escape(title).replace("\n", "<br/>")
        title_block = Paragraph(
            f"<font size='18'><b>{title_html}</b></font><br/><font size='9'>{escape(subtitle)}</font>",
            ParagraphStyle(
                "PdfHeaderTitle",
                fontName="Helvetica",
                fontSize=12,
                leading=22,
                textColor="#082f6f",
                alignment=1,
            ),
        )

        info_width = 4.3 * cm
        info_table = Table(
            [
                [self._p("Formatos: Habittauio", 8, bold=True, align="LEFT", raw=True)],
                [self._p("Fecha de impresión", 8, bold=True, align="LEFT", raw=True)],
                [self._p(self._spanish_date(datetime.now().date()), 8, align="LEFT", raw=True)],
                [self._p(escape(right_text or ""), 8, align="LEFT", raw=True)] if right_text else [self._p("", 8, align="LEFT", raw=True)],
            ],
            colWidths=[info_width],
        )
        info_table.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.8, _PDF_BLUE),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ]))

        header = Table(
            [[logo, title_block, info_table]],
            colWidths=[3.2 * cm, width - 7.5 * cm, info_width],
        )
        header.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 1.2, _PDF_BLUE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (1, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING", (2, 0), (2, 0), 0),
            ("RIGHTPADDING", (2, 0), (2, 0), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ("LINEAFTER", (0, 0), (0, 0), 0.8, _PDF_BLUE),
            ("LINEAFTER", (1, 0), (1, 0), 0.8, _PDF_BLUE),
        ]))
        return [header, Spacer(1, 0.28 * cm)]

    def _metric_cards(self, cards: list[tuple[str, str, str]], width: float) -> Table:
        row = []
        col_width = width / len(cards)
        for title, value, icon in cards:
            card = Table(
                [[self._p(title, 9, bold=True, color="#ffffff")], [self._p(value, 18, bold=True, color="#072f6e")]],
                colWidths=[col_width - 0.25 * cm],
                rowHeights=[0.65 * cm, 1.45 * cm],
            )
            card.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.8, _PDF_BORDER),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))
            row.append(card)
        table = Table([row], colWidths=[col_width] * len(cards))
        table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
        return table

    def _section_title(self, title: str, width: float) -> Table:
        table = Table([[self._p(title, 14, bold=True, color="#07316d", align="LEFT"), ""]], colWidths=[6 * cm, width - 6 * cm])
        table.setStyle(TableStyle([
            ("LINEAFTER", (0, 0), (0, 0), 0, colors.white),
            ("LINEBELOW", (1, 0), (1, 0), 1.2, _PDF_BLUE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return table

    def _report_footer(self, width: float, building: Optional[dict] = None) -> list:
        line = HRFlowable(width=width * 0.52, thickness=1, color=_PDF_BLUE, hAlign="CENTER")
        return [
            Spacer(1, 0.22 * cm),
            line,
            self._p("La Administración", 12, bold=True),
            self._p("Período 2026-2027", 10),
            Spacer(1, 0.18 * cm),
            build_pdf_footer_bar(building or {}, width=width),
        ]

    def _signature_grid(
        self,
        width: float,
        building: Optional[dict],
        document_tag: str,
        *,
        signer_name: str,
        signer_role: str,
        file_name: str | None = None,
    ) -> Table:
        qr_value = f"{document_tag}|{datetime.now().strftime('%Y%m%d%H%M%S')}|{get_building_name(building)}"
        return build_pdf_signature_seal_qr_grid(
            building or {},
            width=width,
            qr_value=qr_value,
            signer_name=signer_name,
            signer_role=signer_role,
            file_name=file_name,
        )

    def _append_signature_grid(
        self,
        story: list,
        *,
        width: float,
        building: Optional[dict],
        document_tag: str,
        signer_name: str = "Usuario del sistema",
        signer_role: str = "Rol no definido",
        file_name: str | None = None,
    ) -> None:
        story.extend(
            [
                Spacer(1, 0.24 * cm),
                self._signature_grid(
                    width,
                    building,
                    document_tag,
                    signer_name=signer_name,
                    signer_role=signer_role,
                    file_name=file_name,
                ),
            ]
        )

    def _footer_callback(self, building: Optional[dict], width: float):
        def draw_footer(canvas, doc):
            canvas.saveState()
            footer = build_pdf_footer_bar(building or {}, width=doc.width, page_text=f"Página {doc.page}")
            _, footer_height = footer.wrap(doc.width, doc.bottomMargin)
            footer.drawOn(canvas, doc.leftMargin, doc.bottomMargin - footer_height)
            canvas.restoreState()

        return draw_footer

    def _styled_table(self, data: list[list], col_widths: list[float], *, font_size: int = 7, total_rows: Optional[list[int]] = None) -> Table:
        table = Table(data, colWidths=col_widths, repeatRows=1)
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), font_size + 1),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("ALIGN", (0, 1), (-1, -1), "LEFT"),
            ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.45, _PDF_BORDER),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f9fd")]),
            ("FONTSIZE", (0, 1), (-1, -1), font_size),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]
        for row in total_rows or []:
            style.extend([
                ("BACKGROUND", (0, row), (-1, row), _PDF_PALE_BLUE),
                ("FONTNAME", (0, row), (-1, row), "Helvetica-Bold"),
                ("TEXTCOLOR", (0, row), (-1, row), _PDF_BLUE),
            ])
        table.setStyle(TableStyle(style))
        return table

    def _comparison_table(self, rows: list[dict], width: float) -> Table:
        max_amount = max([float(row.get("current") or 0) for row in rows] + [1.0])
        data = [[
            self._p("Concepto", 8, bold=True, color="#ffffff", align="LEFT"),
            self._p("Mes actual", 8, bold=True, color="#ffffff"),
            self._p("Mes anterior", 8, bold=True, color="#ffffff"),
            self._p("Valores (USD)", 8, bold=True, color="#ffffff"),
        ]]
        for row in rows:
            current = float(row.get("current") or 0)
            previous = float(row.get("previous") or 0)
            current_bar = "█" * max(1, int((current / max_amount) * 28))
            previous_bar = "█" * max(1, int((previous / max_amount) * 28)) if previous else ""
            data.append([
                self._p(row["label"], 8, bold=True, align="LEFT"),
                self._p(current_bar, 9, color="#003b82"),
                self._p(previous_bar, 9, color="#b8d6f2"),
                self._p(self._money(current), 8, bold=True),
            ])
        return self._styled_table(data, [width * 0.26, width * 0.34, width * 0.25, width * 0.15], font_size=7)

    def _date_label(
        self,
        period: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> str:
        if start_date and end_date:
            return f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        if start_date:
            return f"Desde {start_date.strftime('%d/%m/%Y')}"
        if end_date:
            return f"Hasta {end_date.strftime('%d/%m/%Y')}"
        if period:
            return period
        return "Todos"

    def _table_style(self, font_size: int = 8) -> TableStyle:
        return TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("ALIGN", (0, 1), (-1, -1), "LEFT"),
            ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), font_size + 1),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
            ("BACKGROUND", (0, 1), (-1, -1), _PDF_LIGHT),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("FONTSIZE", (0, 1), (-1, -1), font_size),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ])

    def _style_excel_header(self, ws, headers: list[str]) -> None:
        header_fill = PatternFill(start_color="123C7A", end_color="123C7A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(bottom=Side(style="thin", color="CBD5E1"))
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    async def _payments_report_rows(
        self,
        period: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        status: Optional[str] = None,
    ) -> list[dict]:
        conditions: list[str] = []
        params: list = []
        idx = 1
        if period:
            conditions.append(f"TO_CHAR(p.paid_at, 'YYYY-MM') = ${idx}")
            params.append(period)
            idx += 1
        if status:
            conditions.append(f"p.status = ${idx}")
            params.append(status)
            idx += 1
        if start_date:
            conditions.append(f"p.paid_at >= ${idx}")
            params.append(start_date)
            idx += 1
        if end_date:
            conditions.append(f"p.paid_at <= ${idx}")
            params.append(end_date)

        where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
        rows = await self._payment_repo._conn.fetch(
            f"""
            SELECT p.*, a.code AS apartment_code, a.tower AS apartment_tower, o.full_name AS owner_name
            FROM payments p
            JOIN apartments a ON p.apartment_id = a.id
            JOIN owners o ON p.owner_id = o.id
            {where}
            ORDER BY p.paid_at DESC, p.created_at DESC
            """,
            *params,
        )
        return [dict(row) for row in rows]

    async def _pdf_header(self, title: str, subtitle: str, width: float) -> list:
        conn = getattr(self._payment_repo, "_conn", None)
        building = await get_default_building_config(conn)
        return build_pdf_brand_header(
            title,
            subtitle,
            building,
            width=width,
        )

    async def _payments(
        self,
        period: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> list[dict]:
        conditions = ["p.status IN ('REGISTRADO', 'APROBADO')"]
        params: list = []
        idx = 1
        if period:
            conditions.append(f"TO_CHAR(p.paid_at, 'YYYY-MM') = ${idx}")
            params.append(period)
            idx += 1
        if start_date:
            conditions.append(f"p.paid_at >= ${idx}")
            params.append(start_date)
            idx += 1
        if end_date:
            conditions.append(f"p.paid_at <= ${idx}")
            params.append(end_date)
            idx += 1

        rows = await self._payment_repo._conn.fetch(
            f"""
            SELECT p.*, a.code AS apartment_code, a.tower AS apartment_tower, o.full_name AS owner_name
            FROM payments p
            JOIN apartments a ON p.apartment_id = a.id
            JOIN owners o ON p.owner_id = o.id
            WHERE {' AND '.join(conditions)}
            ORDER BY p.paid_at DESC, p.created_at DESC
            """,
            *params,
        )
        return [dict(row) for row in rows]

    async def _incomes(
        self,
        period: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> list[dict]:
        if not self._income_repo:
            return []
        return await self._income_repo.get_all(
            period=period,
            status="REGISTRADO",
            start_date=start_date,
            end_date=end_date,
        )

    async def _expenses(
        self,
        period: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> list[dict]:
        if period and not start_date and not end_date:
            return await self._monthly_expenses(period)
        if not start_date and not end_date:
            expenses_payload = await self._expense_repo.get_by_month(period)
            expenses = expenses_payload.get("data", []) if isinstance(expenses_payload, dict) else expenses_payload
            return [expense for expense in expenses if self._is_valid_expense(expense)]

        conditions = ["status NOT IN ('ANULADO', 'ANULADA')"]
        params: list = []
        idx = 1
        if period:
            conditions.append(f"TO_CHAR(date, 'YYYY-MM') = ${idx}")
            params.append(period)
            idx += 1
        if start_date:
            conditions.append(f"date >= ${idx}")
            params.append(start_date)
            idx += 1
        if end_date:
            conditions.append(f"date <= ${idx}")
            params.append(end_date)

        rows = await self._expense_repo._conn.fetch(
            f"""
            SELECT *
            FROM expenses
            WHERE {' AND '.join(conditions)}
            ORDER BY date DESC, created_at DESC
            """,
            *params,
        )
        return [dict(row) for row in rows]

    async def monthly_balance_summary(self, period: Optional[str] = None) -> dict:
        target_period = self._validate_month_period(period or date.today().strftime("%Y-%m"))
        previous_period = self._previous_period(target_period)

        payments = await self._monthly_payments(target_period)
        incomes = await self._monthly_incomes(target_period)
        expenses = await self._monthly_expenses(target_period)
        previous_payments = await self._monthly_payments(previous_period)
        previous_incomes = await self._monthly_incomes(previous_period)
        previous_expenses = await self._monthly_expenses(previous_period)

        income_rows = [*payments, *incomes]
        previous_income_rows = [*previous_payments, *previous_incomes]

        income_total = self._sum_amount(income_rows)
        expense_total = self._sum_amount(expenses)
        net_balance = income_total - expense_total

        previous_income_total = self._sum_amount(previous_income_rows)
        previous_expense_total = self._sum_amount(previous_expenses)
        previous_net_balance = previous_income_total - previous_expense_total

        return {
            "period": target_period,
            "income_total": income_total,
            "expense_total": expense_total,
            "net_balance": net_balance,
            "income_breakdown": self._build_breakdown(
                [self._normalize_payment_income(row) for row in payments]
                + [self._normalize_manual_income(row) for row in incomes],
                "income_category",
                "Otros ingresos",
            ),
            "expense_breakdown": self._build_breakdown(expenses, "category", "Sin categoría"),
            "previous_period_variation": {
                "income_pct": self._variation_percent(income_total, previous_income_total),
                "expense_pct": self._variation_percent(expense_total, previous_expense_total),
                "net_balance_pct": self._variation_percent(net_balance, previous_net_balance),
            },
        }

    async def dashboard_stats(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> dict:
        payments = await self._payments(start_date=start_date, end_date=end_date)
        incomes = await self._incomes(start_date=start_date, end_date=end_date)
        expenses = await self._expenses(start_date=start_date, end_date=end_date)
        delinquency = await self._delinquency.get_stats()

        total_revenue = sum(Decimal(str(p.get("amount", 0))) for p in [*payments, *incomes])
        total_expenses = sum(Decimal(str(e.get("amount", 0))) for e in expenses)
        net_income = total_revenue - total_expenses

        previous_revenue = Decimal("0")
        previous_expenses = Decimal("0")
        if start_date and end_date:
            days = (end_date - start_date).days + 1
            previous_end = start_date - timedelta(days=1)
            previous_start = previous_end - timedelta(days=days - 1)
            prev_payments = await self._payments(
                start_date=previous_start,
                end_date=previous_end,
            )
            prev_incomes = await self._incomes(
                start_date=previous_start,
                end_date=previous_end,
            )
            prev_expenses = await self._expenses(
                start_date=previous_start,
                end_date=previous_end,
            )
            previous_revenue = sum(Decimal(str(p.get("amount", 0))) for p in [*prev_payments, *prev_incomes])
            previous_expenses = sum(Decimal(str(e.get("amount", 0))) for e in prev_expenses)

        def change(current: Decimal, previous: Decimal) -> Optional[float]:
            if previous <= 0:
                return None
            return round(float((current - previous) / previous * 100), 2)

        category_totals: dict[str, Decimal] = {}
        for expense in expenses:
            category = expense.get("category") or "Sin categoría"
            category_totals[category] = category_totals.get(category, Decimal("0")) + Decimal(
                str(expense.get("amount", 0))
            )

        start_period = start_date.strftime("%Y-%m") if start_date else None
        end_period = end_date.strftime("%Y-%m") if end_date else None
        period_conditions = []
        period_params = []
        if start_period:
            period_conditions.append(f"period >= ${len(period_params) + 1}")
            period_params.append(start_period)
        if end_period:
            period_conditions.append(f"period <= ${len(period_params) + 1}")
            period_params.append(end_period)
        period_where = f"WHERE {' AND '.join(period_conditions)}" if period_conditions else ""
        fee_rows = await self._payment_repo._conn.fetch(
            f"""
            SELECT period, COALESCE(SUM(amount), 0) AS total
            FROM apartment_fees
            {period_where}
            GROUP BY period
            ORDER BY period ASC
            """,
            *period_params,
        )

        expected_by_period = {row["period"]: float(row["total"]) for row in fee_rows}
        collected_by_period: dict[str, float] = {}
        for payment in payments:
            period_key = payment.get("period") or (
                payment.get("paid_at").strftime("%Y-%m") if payment.get("paid_at") else ""
            )
            if not period_key:
                continue
            collected_by_period[period_key] = collected_by_period.get(period_key, 0.0) + float(
                payment.get("amount", 0)
            )

        periods = sorted(set(expected_by_period) | set(collected_by_period))
        monthly = [
            {
                "period": period_key,
                "expected": expected_by_period.get(period_key, 0.0),
                "collected": collected_by_period.get(period_key, 0.0),
            }
            for period_key in periods[-6:]
        ]

        owner_count = await self._payment_repo._conn.fetchval(
            "SELECT COUNT(*) FROM owners WHERE status = 'ACTIVO'"
        )
        apartment_count = await self._payment_repo._conn.fetchval(
            "SELECT COUNT(*) FROM apartments WHERE status = 'ACTIVO'"
        )

        arrears = []
        for unit in delinquency.get("units", [])[:8]:
            if unit.get("90_plus_days", 0) > 0:
                days_overdue = "90+ Days"
                risk_level = "High"
            elif unit.get("60_days", 0) > 0:
                days_overdue = "60 Days"
                risk_level = "Medium"
            else:
                days_overdue = "30 Days"
                risk_level = "Low"
            arrears.append({
                "unit": unit.get("unit"),
                "owner": unit.get("owner_name"),
                "email": unit.get("email"),
                "amount_due": unit.get("total_debt", 0),
                "days_overdue": days_overdue,
                "risk_level": risk_level,
            })

        fee_details = await self._fees_report_rows(start_date, end_date)
        expense_details = sorted(
            expenses,
            key=lambda row: row.get("date") or date.min,
        )

        return {
            "summary": {
                "total_revenue": float(total_revenue),
                "total_expenses": float(total_expenses),
                "net_income": float(net_income),
                "revenue_change_percent": change(total_revenue, previous_revenue),
                "expense_change_percent": change(total_expenses, previous_expenses),
                "net_income_change_percent": change(
                    net_income, previous_revenue - previous_expenses
                ),
            },
            "expense_categories": [
                {"category": category, "amount": float(amount)}
                for category, amount in sorted(
                    category_totals.items(), key=lambda item: item[1], reverse=True
                )
            ],
            "monthly": monthly,
            "arrears": arrears,
            "fee_details": [
                {
                    "period": row.get("period"),
                    "apartment_code": row.get("apartment_code"),
                    "owner_name": row.get("owner_name"),
                    "amount": float(row.get("amount") or 0),
                    "paid_amount": float(row.get("paid_amount") or 0),
                    "pending_amount": float(row.get("pending_amount") or 0),
                    "status": row.get("status"),
                }
                for row in fee_details[:8]
            ],
            "expense_details": [
                {
                    "date": row.get("date"),
                    "concept": row.get("concept") or row.get("description") or "",
                    "category": row.get("category") or "Sin categoría",
                    "amount": float(row.get("amount") or 0),
                }
                for row in expense_details[:8]
            ],
            "risk_summary": {
                "high": sum(1 for row in arrears if row["risk_level"] == "High"),
                "medium": sum(1 for row in arrears if row["risk_level"] == "Medium"),
                "low": sum(1 for row in arrears if row["risk_level"] == "Low"),
            },
            "system": {
                "active_owners": int(owner_count or 0),
                "active_apartments": int(apartment_count or 0),
                "delinquent_units": delinquency.get("summary", {}).get("delinquent_units", 0),
            },
        }

    async def delinquency_csv(self) -> bytes:
        owners = await self._delinquency.list_owners()
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(
            ["Propietario", "Email", "Documento", "Deuda Total", "Períodos Vencidos", "Estado"]
        )
        for o in owners:
            writer.writerow(
                [
                    o["owner_name"],
                    o.get("email") or "",
                    o["document_id"],
                    o["deuda_total"],
                    o["periodos_vencidos"],
                    self._delinquency_status_label(o["status"]),
                ]
            )
        return output.getvalue().encode("utf-8-sig")

    async def income_csv(
        self,
        period: Optional[str],
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        payments = await self._income_entries(period, start_date, end_date)
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(
            ["Fecha", "Origen", "Concepto", "Propietario", "Departamento", "Período", "Monto", "Método", "Estado"]
        )
        for p in payments:
            writer.writerow(
                [
                    p.get("income_date") or p.get("paid_at", ""),
                    p.get("income_label", ""),
                    p.get("income_concept", ""),
                    p.get("owner_name", ""),
                    p.get("apartment_code", ""),
                    p.get("period", ""),
                    p.get("amount", 0),
                    p.get("method", ""),
                    p.get("status", ""),
                ]
            )
        return output.getvalue().encode("utf-8-sig")

    async def balance_csv(
        self,
        period: Optional[str],
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        payments = await self._income_entries(period, start_date, end_date)
        expenses = await self._expenses(period, start_date, end_date)

        total_income = sum(Decimal(str(p.get("amount", 0))) for p in payments)
        total_expenses = sum(Decimal(str(e.get("amount", 0))) for e in expenses)
        balance = total_income - total_expenses

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Concepto", "Monto"])
        writer.writerow(["Ingresos consolidados", str(total_income)])
        writer.writerow(["Egresos (gastos)", str(total_expenses)])
        writer.writerow(["Balance neto", str(balance)])
        return output.getvalue().encode("utf-8-sig")

    # ─── PDF REPORTS ─────────────────────────────────────────────────────────

    def _draw_double_horizontal_bar(self, val_curr: float, val_prev: float, max_val: float, width: float) -> Drawing:
        d = Drawing(width, 24)
        # Background track for current month
        d.add(Rect(0, 14, width, 6, fillColor=colors.HexColor("#f1f5f9"), strokeColor=None))
        curr_width = (val_curr / max_val) * width if max_val > 0 else 0
        curr_width = min(max(curr_width, 0), width)
        if curr_width > 0:
            d.add(Rect(0, 14, curr_width, 6, fillColor=colors.HexColor("#0b3c7d"), strokeColor=None))
            
        # Background track for previous month
        d.add(Rect(0, 4, width, 6, fillColor=colors.HexColor("#f1f5f9"), strokeColor=None))
        prev_width = (val_prev / max_val) * width if max_val > 0 else 0
        prev_width = min(max(prev_width, 0), width)
        if prev_width > 0:
            d.add(Rect(0, 4, prev_width, 6, fillColor=colors.HexColor("#b9cdfb"), strokeColor=None))
            
        return d

    def _build_income_metric_card(self, title: str, value: str, icon_char: str, card_width: float) -> Table:
        icon_drawing = Drawing(24, 24)
        icon_drawing.add(Circle(12, 12, 11, fillColor=colors.HexColor("#e2e8f0"), strokeColor=colors.HexColor("#cbd5e1"), strokeWidth=0.5))
        icon_drawing.add(String(12, 8.5, icon_char, fontName="Helvetica-Bold", fontSize=10, fillColor=colors.HexColor("#0b3c7d"), textAnchor="middle"))
        
        # Calculate dynamic font size based on value length to prevent overflow/wrap
        font_size = 11.0
        if len(value) > 16:
            font_size = 7.5
        elif len(value) > 12:
            font_size = 9.0

        card_data = [
            [self._p(title.upper(), 6.5, bold=True, color="#ffffff", align="CENTER")],
            [Table([
                [icon_drawing, self._p(value, font_size, bold=True, color="#0b3c7d", align="LEFT")]
            ], colWidths=[28, card_width - 38], style=[
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ])]
        ]
        
        card_table = Table(card_data, colWidths=[card_width])
        card_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        return card_table

    def _stacked_monto(self, curr_txt: str, prev_txt: str, cell_width: float) -> Table:
        t = Table([
            [self._p(curr_txt, 7.5, bold=True, color="#0b3c7d", align="RIGHT")],
            [self._p(prev_txt, 7, color="#64748b", align="RIGHT")]
        ], colWidths=[cell_width])
        t.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ]))
        return t

    def _build_legend(self, curr_name: str, prev_name: str) -> Table:
        box_curr = Drawing(8, 8)
        box_curr.add(Rect(0, 0, 8, 8, fillColor=colors.HexColor("#0b3c7d"), strokeColor=None))
        box_prev = Drawing(8, 8)
        box_prev.add(Rect(0, 0, 8, 8, fillColor=colors.HexColor("#b9cdfb"), strokeColor=None))
        
        legend_table = Table([
            [box_curr, self._p(f"Mes actual ({curr_name})", 7.5, bold=True, color="#475569"),
             box_prev, self._p(f"Mes anterior ({prev_name})", 7.5, bold=True, color="#475569")]
        ], colWidths=[12, 140, 12, 140])
        legend_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
        ]))
        return legend_table

    async def income_pdf(
        self,
        period: Optional[str],
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        compare_period: Optional[str] = None,
    ) -> bytes:
        current_period = period or (start_date.strftime("%Y-%m") if start_date else date.today().strftime("%Y-%m"))
        current_start, current_end = self._get_period_dates(current_period)
        current_start = start_date or current_start
        current_end = end_date or current_end
        
        compare_period = compare_period or self._previous_period(current_period)
        compare_start, compare_end = self._get_period_dates(compare_period)
        
        # Fetch current and comparison data
        payments = await self._income_entries(current_period, current_start, current_end)
        compare_payments = await self._income_entries(compare_period, compare_start, compare_end)
        
        total_current = sum(Decimal(str(p.get("amount", 0))) for p in payments)
        payments_total_current = sum(Decimal(str(p.get("amount", 0))) for p in payments if p.get("income_source_type") == "payment")
        manual_total_current = total_current - payments_total_current
        
        total_compare = sum(Decimal(str(p.get("amount", 0))) for p in compare_payments)
        payments_total_compare = sum(Decimal(str(p.get("amount", 0))) for p in compare_payments if p.get("income_source_type") == "payment")
        manual_total_compare = total_compare - payments_total_compare

        output = io.BytesIO()
        width = A4[0] - 2.2 * cm
        doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=1.1 * cm, rightMargin=1.1 * cm, topMargin=0.8 * cm, bottomMargin=0.7 * cm)
        story = []

        building = await get_default_building_config(getattr(self._payment_repo, "_conn", None))
        
        # Header
        story.extend(
            await self._three_column_report_header(
                "Reporte Detallado de Ingresos",
                f"Rango: {self._date_label(period, start_date, end_date)} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                width,
                building=building,
                right_text=f"Período: {self._date_label(period, start_date, end_date)}",
            )
        )
        
        # 1. 3 metric cards at the top
        departments = len({p.get("apartment_code") for p in payments if p.get("apartment_code")})
        col_w_card = (width / 3) - 5
        cards = [
            self._build_income_metric_card("Departamentos", str(departments), "🏢", col_w_card),
            self._build_income_metric_card("Ingresos por alícuotas", self._usd(payments_total_current), "$", col_w_card),
            self._build_income_metric_card("Otros ingresos", self._usd(manual_total_current), "🎁", col_w_card),
        ]
        cards_table = Table([cards], colWidths=[width / 3] * 3)
        cards_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(cards_table)
        story.append(Spacer(1, 0.25 * cm))
        
        # 2. Custom Title with List Icon
        list_icon = Drawing(16, 16)
        list_icon.add(Circle(8, 8, 7.5, fillColor=colors.HexColor("#0b3c7d"), strokeColor=None))
        list_icon.add(Line(5, 10, 11, 10, strokeColor=colors.white, strokeWidth=1))
        list_icon.add(Line(5, 8, 11, 8, strokeColor=colors.white, strokeWidth=1))
        list_icon.add(Line(5, 6, 11, 6, strokeColor=colors.white, strokeWidth=1))
        
        title_table = Table([[list_icon, self._p("Detalle de ingresos", 11, bold=True, color="#07316d", align="LEFT"), ""]], colWidths=[20, 200, width - 220])
        title_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW", (2, 0), (2, 0), 1.2, _PDF_BLUE),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(title_table)
        story.append(Spacer(1, 0.15 * cm))
        
        # 3. Detalle de ingresos table (Grouped by Torre)
        grouped_entries = {}
        for p in payments:
            is_payment = p.get("income_source_type") == "payment"
            tower = p.get("apartment_tower") or ""
            if not is_payment or not tower:
                group_name = "General"
            else:
                group_name = f"TORRE {tower}"
            
            if group_name not in grouped_entries:
                grouped_entries[group_name] = []
            grouped_entries[group_name].append(p)
            
        sorted_groups = sorted([g for g in grouped_entries.keys() if g != "General"])
        if "General" in grouped_entries:
            sorted_groups.append("General")
            
        data = [[
            self._p("Torre", 7, bold=True, color="#ffffff", align="CENTER"),
            self._p("Departamento", 7, bold=True, color="#ffffff", align="CENTER"),
            self._p("Concepto", 7, bold=True, color="#ffffff", align="CENTER"),
            self._p("Monto", 7, bold=True, color="#ffffff", align="RIGHT"),
            self._p("Observación", 7, bold=True, color="#ffffff"),
        ]]
        span_commands = []
        current_row_idx = 1
        
        for g_name in sorted_groups:
            group_rows = grouped_entries[g_name]
            group_rows = sorted(group_rows, key=lambda x: x.get("apartment_code") or "")
            
            start_row = current_row_idx
            end_row = start_row + len(group_rows) - 1
            
            for p in group_rows:
                apt_code = p.get("apartment_code") or ""
                concept = p.get("income_concept") or ""
                obs = p.get("income_label") or "Ingreso registrado"
                if g_name == "General":
                    apt_code = p.get("method") or "Otros ingresos"
                    concept = p.get("income_concept") or "Otros ingresos"
                    obs = p.get("income_label") or "Ingresos adicionales"
                
                data.append([
                    g_name,
                    self._p(apt_code, 7),
                    self._p(concept, 7),
                    self._p(self._usd(p.get("amount", 0)), 7, color="#1e293b", align="RIGHT"),
                    self._p(obs, 7)
                ])
                current_row_idx += 1
                
            if end_row > start_row:
                span_commands.append(("SPAN", (0, start_row), (0, end_row)))
                
        if len(data) == 1:
            data.append([
                self._p("General", 7, align="CENTER"),
                self._p("-", 7, align="CENTER"),
                self._p("Sin ingresos registrados", 7),
                self._p(self._usd(0), 7, color="#1e293b", align="RIGHT"),
                self._p("-", 7),
            ])
            
        table_style_commands = [
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ("ALIGN", (2, 0), (2, -1), "LEFT"),
            ("ALIGN", (3, 0), (3, -1), "RIGHT"),
            ("ALIGN", (4, 0), (4, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("ROWBACKGROUNDS", (1, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ] + span_commands
        
        detail_col_widths = [width * 0.18, width * 0.20, width * 0.26, width * 0.18, width * 0.18]
        story.append(Table(data, colWidths=detail_col_widths, style=TableStyle(table_style_commands)))
        story.append(Spacer(1, 0.18 * cm))
        
        # 4. v/p Alícuotas sum box
        val_alicuotas_box = Table([[self._p("v/p Alícuotas de departamentos suma", 8, bold=True, color="#0b3c7d"), self._p(self._usd(payments_total_current), 9, bold=True, color="#0b3c7d", align="RIGHT")]], colWidths=[width * 0.75, width * 0.25])
        val_alicuotas_box.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(val_alicuotas_box)
        story.append(Spacer(1, 0.22 * cm))
        
        # 5. Section: Comparativo respecto al mes anterior
        chart_icon = Drawing(16, 16)
        chart_icon.add(Circle(8, 8, 7.5, fillColor=colors.HexColor("#0b3c7d"), strokeColor=None))
        chart_icon.add(Rect(5, 4.5, 1.5, 4, fillColor=colors.white, strokeColor=None))
        chart_icon.add(Rect(7.25, 4.5, 1.5, 7, fillColor=colors.white, strokeColor=None))
        chart_icon.add(Rect(9.5, 4.5, 1.5, 5, fillColor=colors.white, strokeColor=None))
        
        comp_title_table = Table([[chart_icon, self._p("Comparativo respecto al mes anterior", 11, bold=True, color="#07316d", align="LEFT"), ""]], colWidths=[20, 240, width - 260])
        comp_title_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW", (2, 0), (2, 0), 1.2, _PDF_BLUE),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(comp_title_table)
        story.append(Spacer(1, 0.15 * cm))
        
        # Legend Row
        curr_period_name = self._period_name(current_period)
        prev_period_name = self._period_name(compare_period)
        legend_bar = self._build_legend(curr_period_name, prev_period_name)
        story.append(legend_bar)
        story.append(Spacer(1, 0.1 * cm))
        
        # Comparison Table
        max_val = max(total_current, total_compare, Decimal("1"))
        
        col_comp_w0 = width * 0.25
        col_comp_w1 = width * 0.55
        col_comp_w2 = width * 0.20
        
        comp_data = [
            [
                "", 
                "", 
                self._p("Monto", 7.5, bold=True, color="#ffffff", align="CENTER")
            ],
            [
                self._p("Ingresos por alícuotas", 7.5, bold=True, color="#1e293b"),
                self._draw_double_horizontal_bar(float(payments_total_current), float(payments_total_compare), float(max_val), col_comp_w1 - 12),
                self._stacked_monto(self._usd(payments_total_current), self._usd(payments_total_compare), col_comp_w2 - 12)
            ],
            [
                self._p("Otros ingresos", 7.5, bold=True, color="#1e293b"),
                self._draw_double_horizontal_bar(float(manual_total_current), float(manual_total_compare), float(max_val), col_comp_w1 - 12),
                self._stacked_monto(self._usd(manual_total_current), self._usd(manual_total_compare), col_comp_w2 - 12)
            ],
            [
                self._p("Total ingresos", 7.5, bold=True, color="#1e293b"),
                self._draw_double_horizontal_bar(float(total_current), float(total_compare), float(max_val), col_comp_w1 - 12),
                self._stacked_monto(self._usd(total_current), self._usd(total_compare), col_comp_w2 - 12)
            ]
        ]
        
        comp_table = Table(comp_data, colWidths=[col_comp_w0, col_comp_w1, col_comp_w2])
        comp_table.setStyle(TableStyle([
            ("BACKGROUND", (2, 0), (2, 0), _PDF_NAVY),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("LINEAFTER", (1, 0), (1, -1), 0.5, _PDF_BORDER),
            ("BOX", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]))
        story.append(comp_table)
        story.append(Spacer(1, 0.15 * cm))
        
        # 6. Full-width total registrados bar
        total_bar = Table([[self._p("Ingresos totales registrados en el período", 8, bold=True, color="#ffffff"), self._p(self._usd(total_current), 9, bold=True, color="#ffffff", align="RIGHT")]], colWidths=[width * 0.75, width * 0.25])
        total_bar.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), _PDF_NAVY),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(total_bar)
        
        self._append_signature_grid(story, width=width, building=building, document_tag="REPORTE-INGRESOS", signer_name="La Administración", signer_role="Período 2026-2027")
        footer = self._footer_callback(building, width)
        doc.build(story, onFirstPage=footer, onLaterPages=footer)
        return output.getvalue()

    def _draw_horizontal_bar(self, val: float, max_val: float, width: float, is_percent: bool = False, color="#1155d9") -> Drawing:
        def format_short_val(v: float) -> str:
            v = float(v)
            if abs(v) >= 1_000_000:
                return f"{v / 1_000_000:.1f}M".replace(".0", "")
            elif abs(v) >= 1_000:
                return f"{v / 1_000:.1f}k".replace(".0", "")
            else:
                return f"{int(v)}"

        d = Drawing(width, 22)
        d.add(Rect(0, 10, width, 5, fillColor=colors.HexColor("#e2e8f0"), strokeColor=None))
        fill_width = (val / max_val) * width if max_val > 0 else 0
        fill_width = min(max(fill_width, 0), width)
        if fill_width > 0:
            d.add(Rect(0, 10, fill_width, 5, fillColor=colors.HexColor(color), strokeColor=None))
        for pct in [0.0, 0.25, 0.50, 0.75, 1.0]:
            x = pct * width
            d.add(Line(x, 10, x, 7, strokeColor=colors.HexColor("#cbd5e1"), strokeWidth=0.5))
            tick_val = pct * max_val
            if is_percent:
                lbl = f"{int(tick_val)}%"
            else:
                lbl = format_short_val(tick_val)
            d.add(String(x, 0, lbl, fontName="Helvetica", fontSize=5.5, fillColor=colors.HexColor("#64748b"), textAnchor="middle"))
        return d

    def _build_metric_card(self, title: str, value: str, subtitle: str, icon_char: str, card_width: float) -> Table:
        icon_drawing = Drawing(24, 24)
        icon_drawing.add(Circle(12, 12, 11, fillColor=colors.HexColor("#e2e8f0"), strokeColor=colors.HexColor("#cbd5e1"), strokeWidth=0.5))
        icon_drawing.add(String(12, 8.5, icon_char, fontName="Helvetica-Bold", fontSize=10, fillColor=colors.HexColor("#475569"), textAnchor="middle"))
        
        # Calculate dynamic font size based on value length to prevent overflow/wrap
        font_size = 9.5
        if len(value) > 16:
            font_size = 7.0
        elif len(value) > 12:
            font_size = 8.0

        card_data = [
            [icon_drawing],
            [self._p(title.upper(), 5.5, bold=True, color="#475569", align="CENTER")],
            [self._p(value, font_size, bold=True, color="#0f172a", align="CENTER")],
            [self._p(subtitle, 5.5, color="#64748b", align="CENTER")]
        ]
        
        card_table = Table(card_data, colWidths=[card_width])
        card_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ]))
        return card_table

    def _draw_comparison_table(
        self,
        prev_name: str,
        curr_name: str,
        metrics: list[dict],
        width: float,
    ) -> Table:
        max_currency = max(
            [float(m["current"]) for m in metrics if not m["is_percent"]] +
            [float(m["previous"]) for m in metrics if not m["is_percent"]] +
            [100.0]
        )
        max_val_currency = self._round_to_nice_max(max_currency * 1.1)
        
        data = [
            [
                self._p("COMPARATIVO CON EL MES ANTERIOR", 8, bold=True, color="#ffffff"),
                "",
                "",
                self._p(f"Comparar períodos: {prev_name} - {curr_name}", 8, bold=True, color="#ffffff", align="RIGHT"),
                ""
            ],
            [
                self._p("Indicador", 7.5, bold=True, color="#ffffff"),
                self._p(f"Mes Anterior\n({prev_name})", 7.5, bold=True, color="#ffffff", align="CENTER"),
                self._p(f"Mes Actual\n({curr_name})", 7.5, bold=True, color="#ffffff", align="CENTER"),
                self._p("Variación (USD)", 7.5, bold=True, color="#ffffff", align="RIGHT"),
                self._p("Variación (%)", 7.5, bold=True, color="#ffffff", align="RIGHT")
            ]
        ]
        
        col_w0 = width * 0.22
        col_w1 = width * 0.28
        col_w2 = width * 0.28
        col_w3 = width * 0.11
        col_w4 = width * 0.11
        
        comp_style = [
            ("SPAN", (0, 0), (2, 0)),
            ("SPAN", (3, 0), (4, 0)),
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
            ("BACKGROUND", (0, 1), (-1, 1), _PDF_BLUE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (2, -1), "CENTER"),
            ("ALIGN", (3, 0), (4, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("ROWBACKGROUNDS", (0, 2), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]
        
        for idx, m in enumerate(metrics):
            curr_val = float(m["current"])
            prev_val = float(m["previous"])
            is_percent = m["is_percent"]
            inverse = m["inverse"]
            
            if is_percent:
                curr_txt = f"{curr_val:.2f}%"
                prev_txt = f"{prev_val:.2f}%"
                diff = curr_val - prev_val
                pct_diff = diff
            else:
                curr_txt = self._usd(curr_val)
                prev_txt = self._usd(prev_val)
                diff = curr_val - prev_val
                pct_diff = (diff / prev_val * 100) if prev_val > 0 else 0.0
                
            is_positive_change = diff >= 0
            if is_positive_change:
                sign_str = "+"
                arr_str = "^"
                if inverse:
                    var_color = "#9a3412"
                else:
                    var_color = "#166534"
            else:
                sign_str = ""
                arr_str = "v"
                if inverse:
                    var_color = "#166534"
                else:
                    var_color = "#9a3412"
                    
            if diff == 0:
                sign_str = ""
                arr_str = ""
                var_color = "#475569"
                
            if is_percent:
                var_usd_txt = "-"
                var_pct_txt = f"{sign_str}{pct_diff:.2f}% {arr_str}"
            else:
                var_usd_txt = f"{sign_str}{self._usd(diff)}"
                var_pct_txt = f"{sign_str}{pct_diff:.2f}% {arr_str}"
                
            bar_max = 100.0 if is_percent else max_val_currency
            prev_bar = self._draw_horizontal_bar(prev_val, bar_max, col_w1 - 12, is_percent, color="#94a3b8")
            curr_bar = self._draw_horizontal_bar(curr_val, bar_max, col_w2 - 12, is_percent, color="#1155d9")
            
            row_idx = len(data)
            data.append([
                self._p(m["label"], 7.5, bold=True, color="#1e293b"),
                [self._p(prev_txt, 7.5, bold=True, color="#475569", align="CENTER"), prev_bar],
                [self._p(curr_txt, 7.5, bold=True, color="#0f172a", align="CENTER"), curr_bar],
                self._p(var_usd_txt, 7, bold=True, color=var_color, align="RIGHT"),
                self._p(var_pct_txt, 7, bold=True, color=var_color, align="RIGHT")
            ])
            
        table = Table(data, colWidths=[col_w0, col_w1, col_w2, col_w3, col_w4])
        table.setStyle(TableStyle(comp_style))
        return table

    def _round_to_nice_max(self, val: float) -> float:
        if val <= 100:
            return 100.0
        elif val <= 500:
            return 500.0
        elif val <= 1000:
            return 1000.0
        elif val <= 5000:
            return 5000.0
        elif val <= 10000:
            return 10000.0
        else:
            return math.ceil(val / 5000) * 5000.0

    def _get_period_dates(self, period: str) -> tuple[date, date]:
        try:
            year, month = map(int, period.split("-"))
            last_day = calendar.monthrange(year, month)[1]
            return date(year, month, 1), date(year, month, last_day)
        except Exception:
            today = date.today()
            last_day = calendar.monthrange(today.year, today.month)[1]
            return date(today.year, today.month, 1), date(today.year, today.month, last_day)

    async def balance_pdf(
        self,
        period: Optional[str],
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        compare_period: Optional[str] = None,
    ) -> bytes:
        current_period = period or (start_date.strftime("%Y-%m") if start_date else date.today().strftime("%Y-%m"))
        current_start, current_end = self._get_period_dates(current_period)
        current_start = start_date or current_start
        current_end = end_date or current_end
        
        compare_period = compare_period or self._previous_period(current_period)
        compare_start, compare_end = self._get_period_dates(compare_period)
        
        # Fetch current period data
        current_payments = await self._income_entries(current_period, current_start, current_end)
        current_expenses = await self._expenses(current_period, current_start, current_end)
        current_fee_details = await self._fees_report_rows(current_start, current_end)
        
        # Fetch comparison period data
        compare_payments = await self._income_entries(compare_period, compare_start, compare_end)
        compare_expenses = await self._expenses(compare_period, compare_start, compare_end)
        compare_fee_details = await self._fees_report_rows(compare_start, compare_end)
        
        # Calculations for current period
        total_income_current = sum(Decimal(str(p.get("amount", 0))) for p in current_payments)
        total_expenses_current = sum(Decimal(str(e.get("amount", 0))) for e in current_expenses)
        balance_current = total_income_current - total_expenses_current
        projected_current = sum(Decimal(str(row.get("amount") or 0)) for row in current_fee_details)
        valores_por_recuperar_current = max(Decimal("0"), projected_current - total_income_current)
        saldo_con_valores_por_recuperar_current = balance_current + valores_por_recuperar_current
        efficiency_current = (total_income_current / projected_current * 100) if projected_current > 0 else Decimal("0")
        
        # Calculations for comparison period
        total_income_compare = sum(Decimal(str(p.get("amount", 0))) for p in compare_payments)
        total_expenses_compare = sum(Decimal(str(e.get("amount", 0))) for e in compare_expenses)
        balance_compare = total_income_compare - total_expenses_compare
        projected_compare = sum(Decimal(str(row.get("amount") or 0)) for row in compare_fee_details)
        valores_por_recuperar_compare = max(Decimal("0"), projected_compare - total_income_compare)
        saldo_con_valores_por_recuperar_compare = balance_compare + valores_por_recuperar_compare
        efficiency_compare = (total_income_compare / projected_compare * 100) if projected_compare > 0 else Decimal("0")

        def money_cell(value, *, bold: bool = False, color: str = "#1e293b"):
            return self._p(self._usd(value), 7, bold=bold, color=color, align="RIGHT")
        
        output = io.BytesIO()
        width = A4[0] - 2.2 * cm
        doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=1.1 * cm, rightMargin=1.1 * cm, topMargin=0.8 * cm, bottomMargin=0.7 * cm)
        story = []
        building = await get_default_building_config(getattr(self._payment_repo, "_conn", None))
        
        # Header - original header formatting preserved
        story.extend(
            await self._three_column_report_header(
                "BALANCE DE FIN DE MES\nINGRESOS Y EGRESOS",
                f"Rango: {self._date_label(period, start_date, end_date)} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                width,
                building=building,
                right_text="Moneda: USD - Dólares",
            )
        )
        
        # 1. Metric Cards (6 cards in a row)
        col_w_card = (width / 6) - 5
        cards = [
            self._build_metric_card("Ingreso proyectado por alícuotas", self._usd(projected_current), "Proyección del mes", "P", col_w_card),
            self._build_metric_card("Ingreso efectivo por alícuotas", self._usd(total_income_current), "Recaudado al día de hoy", "$", col_w_card),
            self._build_metric_card("Gastos del mes", self._usd(total_expenses_current), "Total gastos del mes", "▼", col_w_card),
            self._build_metric_card("Saldo (superávit / déficit)", self._usd(balance_current), "Superávit del mes" if balance_current >= 0 else "Déficit del mes", "S", col_w_card),
            self._build_metric_card("Saldo con valores por recuperar", self._usd(saldo_con_valores_por_recuperar_current), "Pendiente por cobrar", "R", col_w_card),
            self._build_metric_card("Eficiencia en recaudación", f"{float(efficiency_current):.2f}%", "Eficiencia este mes", "%", col_w_card),
        ]
        cards_table = Table([cards], colWidths=[width / 6] * 6)
        cards_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(cards_table)
        story.append(Spacer(1, 0.22 * cm))
        
        # Section Title: DETALLE DEL MES
        curr_period_name = self._period_name(current_period).upper()
        detail_title = Table(
            [[self._p(f"DETALLE DEL MES - {curr_period_name}", 14, bold=True, color="#0b3c7d", align="CENTER")]],
            colWidths=[width],
        )
        detail_title.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LINEBELOW", (0, 0), (-1, -1), 0.7, _PDF_BORDER),
        ]))
        story.append(detail_title)
        
        # 2. Side-by-side Resúmenes tables (Section 2)
        income_by_category = self._build_breakdown(current_payments, "income_category", "Otros ingresos")
        alicuotas_amount = Decimal("0")
        otros_amount = Decimal("0")
        for item in income_by_category:
            if "alícuota" in item["label"].lower() or "pago" in item["label"].lower():
                alicuotas_amount += item["amount"]
            else:
                otros_amount += item["amount"]
                
        income_rows = [
            [self._p("RESUMEN DE INGRESOS (POR CATEGORÍA)", 7.5, bold=True, color="#ffffff"), self._p("↗", 8, bold=True, color="#ffffff", align="RIGHT")],
            [self._p("Concepto", 7, bold=True, color="#ffffff"), self._p("Monto (USD)", 7, bold=True, color="#ffffff", align="RIGHT")],
            [self._p("Alícuotas y pagos", 7, color="#1e293b"), money_cell(alicuotas_amount)],
            [self._p("Otros ingresos", 7, color="#1e293b"), money_cell(otros_amount)],
            [self._p("Proyección alícuotas del mes", 7, color="#64748b"), money_cell(projected_current, color="#64748b")],
            [self._p("TOTAL INGRESOS EFECTIVOS", 7, bold=True, color="#123c7a"), money_cell(total_income_current, bold=True, color="#123c7a")]
        ]
        
        income_table = Table(income_rows, colWidths=[width * 0.485 * 0.68, width * 0.485 * 0.32])
        income_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
            ("BACKGROUND", (0, 1), (-1, 1), _PDF_BLUE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("ROWBACKGROUNDS", (0, 2), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e2e8f0")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        
        expense_by_category = self._build_breakdown(current_expenses, "category", "Otros")
        expense_rows = [
            [self._p("RESUMEN DE EGRESOS (POR CATEGORÍA)", 7.5, bold=True, color="#ffffff"), self._p("↓", 8, bold=True, color="#ffffff", align="RIGHT")],
            [self._p("Categoría", 7, bold=True, color="#ffffff"), self._p("Monto (USD)", 7, bold=True, color="#ffffff", align="RIGHT")]
        ]
        for item in expense_by_category[:3]:
            expense_rows.append([self._p(item["label"], 7, color="#1e293b"), money_cell(item["amount"])])
            
        while len(expense_rows) < 5:
            expense_rows.append([self._p("-", 7, color="#64748b"), money_cell(0, color="#64748b")])
            
        expense_rows.append([self._p("TOTAL EGRESOS", 7, bold=True, color="#123c7a"), money_cell(total_expenses_current, bold=True, color="#123c7a")])
        
        expense_table = Table(expense_rows, colWidths=[width * 0.485 * 0.68, width * 0.485 * 0.32])
        expense_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
            ("BACKGROUND", (0, 1), (-1, 1), _PDF_BLUE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("ROWBACKGROUNDS", (0, 2), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e2e8f0")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        
        details = Table([[income_table, "", expense_table]], colWidths=[width * 0.485, width * 0.03, width * 0.485])
        details.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(details)
        story.append(Spacer(1, 0.22 * cm))
        
        # 3. Middle box section side-by-side (Section 4)
        saldo_drawing = Drawing(30, 24)
        saldo_drawing.add(Circle(15, 12, 11, fillColor=colors.HexColor("#e0e7ff"), strokeColor=colors.HexColor("#c7d2fe"), strokeWidth=0.5))
        saldo_drawing.add(String(15, 8.5, "⚖", fontName="Helvetica-Bold", fontSize=11, fillColor=colors.HexColor("#3730a3"), textAnchor="middle"))
        
        card_saldo_data = [
            [saldo_drawing],
            [self._p("SALDO (SUPERÁVIT / DÉFICIT)", 6.5, bold=True, color="#475569", align="CENTER")],
            [self._p(self._usd(balance_current), 9, bold=True, color="#1e3a8a", align="CENTER")],
            [self._p("Superávit del mes" if balance_current >= 0 else "Déficit del mes", 6, color="#64748b", align="CENTER")]
        ]
        card_saldo = Table(card_saldo_data, colWidths=[width * 0.31])
        card_saldo.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        
        recuperar_drawing = Drawing(30, 24)
        recuperar_drawing.add(Circle(15, 12, 11, fillColor=colors.HexColor("#e0e7ff"), strokeColor=colors.HexColor("#c7d2fe"), strokeWidth=0.5))
        recuperar_drawing.add(String(15, 8.5, "R", fontName="Helvetica-Bold", fontSize=11, fillColor=colors.HexColor("#3730a3"), textAnchor="middle"))
        
        card_recuperar_data = [
            [recuperar_drawing],
            [self._p("SALDO CON VALORES POR RECUPERAR", 6.5, bold=True, color="#475569", align="CENTER")],
            [self._p(self._usd(saldo_con_valores_por_recuperar_current), 9, bold=True, color="#1e3a8a", align="CENTER")],
            [self._p("Pendiente por cobrar", 6, color="#64748b", align="CENTER")]
        ]
        card_recuperar = Table(card_recuperar_data, colWidths=[width * 0.31])
        card_recuperar.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        
        formula_rows = [
            [self._p("Ingreso proyectado por alícuotas", 7, color="#334155"), money_cell(projected_current, bold=True)],
            [self._p("(-) Ingreso efectivo por alícuotas", 7, color="#334155"), money_cell(total_income_current, bold=True)],
            [self._p("(-) Gastos del mes", 7, color="#334155"), money_cell(total_expenses_current, bold=True)],
            [self._p("SALDO (SUPERÁVIT / DÉFICIT)", 7.5, bold=True, color="#ffffff"), money_cell(balance_current, bold=True, color="#ffffff")]
        ]
        formula_table = Table(formula_rows, colWidths=[width * 0.32 * 0.60, width * 0.32 * 0.40])
        formula_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("BACKGROUND", (0, 0), (-1, -2), colors.HexColor("#f8fafc")),
            ("BACKGROUND", (0, -1), (-1, -1), _PDF_NAVY),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        
        middle_section = Table([
            [card_saldo, "", card_recuperar, "", formula_table]
        ], colWidths=[width * 0.31, width * 0.03, width * 0.31, width * 0.03, width * 0.32])
        middle_section.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(middle_section)
        story.append(Spacer(1, 0.22 * cm))
        
        # 4. Side-by-side details tables (Section 3)
        alicuotas_detail_rows = [
            [self._p("DETALLE DE ALÍCUOTAS (POR DEPARTAMENTO)", 7.5, bold=True, color="#ffffff"), "", "", self._p("🏢", 8, bold=True, color="#ffffff", align="CENTER")],
            [self._p("Departamento", 7, bold=True, color="#ffffff"), self._p("Propietario", 7, bold=True, color="#ffffff"), self._p("Monto (USD)", 7, bold=True, color="#ffffff", align="RIGHT"), self._p("Estado", 7, bold=True, color="#ffffff", align="CENTER")]
        ]
        
        detail_style = [
            ("SPAN", (0, 0), (2, 0)),
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
            ("BACKGROUND", (0, 1), (-1, 1), _PDF_BLUE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (1, -1), "LEFT"),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ("ALIGN", (3, 0), (3, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("ROWBACKGROUNDS", (0, 2), (-1, -3), [colors.white, colors.HexColor("#f8fafc")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ]
        
        for row in current_fee_details[:8]:
            apt_code = row.get("apartment_code") or ""
            owner = row.get("owner_name") or "Sin propietario"
            amount = row.get("amount") or 0
            status = row.get("status") or "PENDIENTE"
            
            row_idx = len(alicuotas_detail_rows)
            if status == "PAGADA":
                status_p = self._p("Pagado", 6, bold=True, color="#166534", align="CENTER")
                detail_style.append(("BACKGROUND", (3, row_idx), (3, row_idx), colors.HexColor("#dcfce7")))
            else:
                status_p = self._p("Pendiente", 6, bold=True, color="#9a3412", align="CENTER")
                detail_style.append(("BACKGROUND", (3, row_idx), (3, row_idx), colors.HexColor("#fef3c7")))
                
            alicuotas_detail_rows.append([
                self._p(apt_code, 7, color="#1e293b"),
                self._p(owner[:20] + "..." if len(owner) > 20 else owner, 7, color="#1e293b"),
                money_cell(amount),
                status_p
            ])
            
        while len(alicuotas_detail_rows) < 7:
            alicuotas_detail_rows.append(["-", "-", money_cell(0, color="#64748b"), "-"])
            
        total_rec_idx = len(alicuotas_detail_rows)
        alicuotas_detail_rows.append([self._p("TOTAL RECAUDADO", 7, bold=True, color="#123c7a"), "", money_cell(total_income_current, bold=True, color="#123c7a"), ""])
        detail_style.append(("SPAN", (0, total_rec_idx), (1, total_rec_idx)))
        detail_style.append(("BACKGROUND", (0, total_rec_idx), (-1, total_rec_idx), colors.HexColor("#e2e8f0")))
        
        total_proj_idx = len(alicuotas_detail_rows)
        alicuotas_detail_rows.append([self._p("TOTAL PROYECTADO (ALÍCUOTAS)", 7, bold=True, color="#123c7a"), "", money_cell(projected_current, bold=True, color="#123c7a"), ""])
        detail_style.append(("SPAN", (0, total_proj_idx), (1, total_proj_idx)))
        detail_style.append(("BACKGROUND", (0, total_proj_idx), (-1, total_proj_idx), colors.HexColor("#e2e8f0")))
        
        apartments_table = Table(alicuotas_detail_rows, colWidths=[width * 0.485 * 0.16, width * 0.485 * 0.38, width * 0.485 * 0.26, width * 0.485 * 0.20])
        apartments_table.setStyle(TableStyle(detail_style))
        
        gastos_detail_rows = [
            [self._p("DETALLE DE GASTOS (POR MOVIMIENTO)", 7.5, bold=True, color="#ffffff"), "", "", self._p("▤", 8, bold=True, color="#ffffff", align="CENTER")],
            [self._p("Fecha", 7, bold=True, color="#ffffff"), self._p("Concepto", 7, bold=True, color="#ffffff"), self._p("Categoría", 7, bold=True, color="#ffffff"), self._p("Monto (USD)", 7, bold=True, color="#ffffff", align="RIGHT")]
        ]
        
        gastos_style = [
            ("SPAN", (0, 0), (2, 0)),
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
            ("BACKGROUND", (0, 1), (-1, 1), _PDF_BLUE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (2, -1), "LEFT"),
            ("ALIGN", (3, 0), (3, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("ROWBACKGROUNDS", (0, 2), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ]
        
        sorted_expenses = sorted(current_expenses, key=lambda e: e.get("date") or date.min)
        for e in sorted_expenses[:8]:
            dt = e.get("date")
            date_str = dt.strftime("%d/%m/%Y") if hasattr(dt, "strftime") else str(dt or "")
            concept = e.get("concept") or ""
            category = e.get("category") or "Sin categoría"
            amount = e.get("amount") or 0
            
            gastos_detail_rows.append([
                self._p(date_str, 7, color="#1e293b"),
                self._p(concept[:22] + "..." if len(concept) > 22 else concept, 7, color="#1e293b"),
                self._p(category, 7, color="#64748b"),
                money_cell(amount)
            ])
            
        while len(gastos_detail_rows) < len(alicuotas_detail_rows) - 1:
            gastos_detail_rows.append(["-", "-", "-", money_cell(0, color="#64748b")])
            
        total_gastos_idx = len(gastos_detail_rows)
        gastos_detail_rows.append([self._p("TOTAL EGRESOS DEL MES", 7, bold=True, color="#123c7a"), "", "", money_cell(total_expenses_current, bold=True, color="#123c7a")])
        gastos_style.append(("SPAN", (0, total_gastos_idx), (2, total_gastos_idx)))
        gastos_style.append(("BACKGROUND", (0, total_gastos_idx), (-1, total_gastos_idx), colors.HexColor("#e2e8f0")))
        
        expenses_table = Table(gastos_detail_rows, colWidths=[width * 0.485 * 0.16, width * 0.485 * 0.36, width * 0.485 * 0.20, width * 0.485 * 0.28])
        expenses_table.setStyle(TableStyle(gastos_style))
        
        details_apartments_expenses = Table([[apartments_table, "", expenses_table]], colWidths=[width * 0.485, width * 0.03, width * 0.485])
        details_apartments_expenses.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(details_apartments_expenses)
        story.append(Spacer(1, 0.22 * cm))
        
        # 5. Comparison Section (Section 5)
        metrics_list = [
            {"label": "Ingreso efectivo por alícuotas", "current": total_income_current, "previous": total_income_compare, "is_percent": False, "inverse": False},
            {"label": "Gastos del mes", "current": total_expenses_current, "previous": total_expenses_compare, "is_percent": False, "inverse": True},
            {"label": "Resultado del mes (Superávit / Déficit)", "current": balance_current, "previous": balance_compare, "is_percent": False, "inverse": False},
            {"label": "Valores por recuperar", "current": valores_por_recuperar_current, "previous": valores_por_recuperar_compare, "is_percent": False, "inverse": True},
            {"label": "Eficiencia en recaudación de alícuotas", "current": efficiency_current, "previous": efficiency_compare, "is_percent": True, "inverse": False},
        ]
        
        prev_name_label = self._period_name(compare_period)
        curr_name_label = self._period_name(current_period)
        comparison_table = self._draw_comparison_table(prev_name_label, curr_name_label, metrics_list, width)
        story.append(comparison_table)
        
        # Signature block & Footer callbacks - original formatting preserved
        self._append_signature_grid(story, width=width, building=building, document_tag="REPORTE-BALANCE", signer_name="Administración", signer_role="Administrador del edificio")
        footer = self._footer_callback(building, width)
        doc.build(story, onFirstPage=footer, onLaterPages=footer)
        return output.getvalue()

    async def payments_pdf(
        self,
        period: Optional[str],
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        compare_period: Optional[str] = None,
        status: Optional[str] = None,
    ) -> bytes:
        current_period = period or (start_date.strftime("%Y-%m") if start_date else date.today().strftime("%Y-%m"))
        current_start, current_end = self._get_period_dates(current_period)
        current_start = start_date or current_start
        current_end = end_date or current_end
        
        compare_period = compare_period or self._previous_period(current_period)
        compare_start, compare_end = self._get_period_dates(compare_period)
        
        payments = await self._payments_report_rows(current_period, current_start, current_end, status)
        compare_payments = await self._payments_report_rows(compare_period, compare_start, compare_end, status)
        
        total_current = sum(Decimal(str(p.get("amount", 0))) for p in payments)
        total_compare = sum(Decimal(str(p.get("amount", 0))) for p in compare_payments)
        
        by_method_current = self._build_breakdown(payments, "method", "Sin método")

        output = io.BytesIO()
        width = A4[0] - 2.2 * cm
        doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=1.1 * cm, rightMargin=1.1 * cm, topMargin=0.8 * cm, bottomMargin=0.7 * cm)
        story = []
        building = await get_default_building_config(getattr(self._payment_repo, "_conn", None))
        
        # Header
        story.extend(
            await self._three_column_report_header(
                "Reporte Detallado de Pagos",
                f"Rango: {self._date_label(period, start_date, end_date)} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                width,
                building=building,
                right_text=f"Total recaudado: {self._money(total_current)}",
            )
        )
        
        # 1. 3 metric cards at the top
        icon_payments = Drawing(24, 24)
        icon_payments.add(Circle(12, 12, 11, fillColor=colors.HexColor("#e2e8f0"), strokeColor=colors.HexColor("#cbd5e1"), strokeWidth=0.5))
        icon_payments.add(Rect(9, 6, 6, 12, fillColor=colors.white, strokeColor=colors.HexColor("#0b3c7d"), strokeWidth=1))
        icon_payments.add(Line(11, 9, 13, 9, strokeColor=colors.HexColor("#0b3c7d"), strokeWidth=0.5))
        icon_payments.add(Line(11, 11, 13, 11, strokeColor=colors.HexColor("#0b3c7d"), strokeWidth=0.5))
        icon_payments.add(Line(11, 13, 13, 13, strokeColor=colors.HexColor("#0b3c7d"), strokeWidth=0.5))
        
        icon_methods = Drawing(24, 24)
        icon_methods.add(Circle(12, 12, 11, fillColor=colors.HexColor("#e2e8f0"), strokeColor=colors.HexColor("#cbd5e1"), strokeWidth=0.5))
        icon_methods.add(Rect(9, 9, 2.5, 2.5, fillColor=colors.HexColor("#0b3c7d"), strokeColor=None))
        icon_methods.add(Rect(12.5, 9, 2.5, 2.5, fillColor=colors.HexColor("#0b3c7d"), strokeColor=None))
        icon_methods.add(Rect(9, 12.5, 2.5, 2.5, fillColor=colors.HexColor("#0b3c7d"), strokeColor=None))
        icon_methods.add(Rect(12.5, 12.5, 2.5, 2.5, fillColor=colors.HexColor("#0b3c7d"), strokeColor=None))
        
        col_w_card = (width / 3) - 5
        cards = [
            self._build_income_metric_card("Pagos", str(len(payments)), "P", col_w_card),
            self._build_income_metric_card("Total recaudado", self._money(total_current), "$", col_w_card),
            self._build_income_metric_card("Métodos", str(len(by_method_current)), "M", col_w_card),
        ]
        # Override Card 1 & 3 icons with vector shapes
        cards[0] = Table([
            [self._p("PAGOS", 6.5, bold=True, color="#ffffff", align="CENTER")],
            [Table([
                [icon_payments, self._p(str(len(payments)), 11, bold=True, color="#0b3c7d", align="LEFT")]
            ], colWidths=[28, col_w_card - 38], style=[
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ])]
        ], colWidths=[col_w_card])
        cards[0].setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        
        cards[2] = Table([
            [self._p("MÉTODOS", 6.5, bold=True, color="#ffffff", align="CENTER")],
            [Table([
                [icon_methods, self._p(str(len(by_method_current)), 11, bold=True, color="#0b3c7d", align="LEFT")]
            ], colWidths=[28, col_w_card - 38], style=[
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ])]
        ], colWidths=[col_w_card])
        cards[2].setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        
        cards_table = Table([cards], colWidths=[width / 3] * 3)
        cards_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(cards_table)
        story.append(Spacer(1, 0.25 * cm))
        
        # 2. Custom Title with List Icon
        list_icon = Drawing(16, 16)
        list_icon.add(Circle(8, 8, 7.5, fillColor=colors.HexColor("#0b3c7d"), strokeColor=None))
        list_icon.add(Line(5, 10, 11, 10, strokeColor=colors.white, strokeWidth=1))
        list_icon.add(Line(5, 8, 11, 8, strokeColor=colors.white, strokeWidth=1))
        list_icon.add(Line(5, 6, 11, 6, strokeColor=colors.white, strokeWidth=1))
        
        title_table = Table([[list_icon, self._p("Detalle de pagos", 11, bold=True, color="#07316d", align="LEFT"), ""]], colWidths=[20, 200, width - 220])
        title_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW", (2, 0), (2, 0), 1.2, _PDF_BLUE),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(title_table)
        story.append(Spacer(1, 0.15 * cm))
        
        # 3. Detalle de pagos table (Grouped by Torre)
        grouped_entries = {}
        for p in payments:
            tower = p.get("apartment_tower") or ""
            group_name = f"TORRE {tower}" if tower else "General"
            
            if group_name not in grouped_entries:
                grouped_entries[group_name] = []
            grouped_entries[group_name].append(p)
            
        sorted_groups = sorted([g for g in grouped_entries.keys() if g != "General"])
        if "General" in grouped_entries:
            sorted_groups.append("General")
            
        data = [["Torre", "Departamento", "Propietario", "Concepto", "Monto", "Método", "Estado"]]
        span_commands = []
        current_row_idx = 1
        
        for g_name in sorted_groups:
            group_rows = grouped_entries[g_name]
            group_rows = sorted(group_rows, key=lambda x: x.get("apartment_code") or "")
            
            start_row = current_row_idx
            end_row = start_row + len(group_rows) - 1
            
            for p in group_rows:
                apt_code = p.get("apartment_code") or ""
                owner = p.get("owner_name") or ""
                concept = f"Pago {p.get('period') or ''}".strip()
                amount = p.get("amount", 0)
                method = p.get("method") or ""
                status_str = p.get("status") or "REGISTRADO"
                
                row_idx = len(data)
                if status_str == "APROBADO":
                    status_p = self._p("Aprobado", 6, bold=True, color="#166534", align="CENTER")
                    table_style_commands_bg = ("BACKGROUND", (6, row_idx), (6, row_idx), colors.HexColor("#dcfce7"))
                else:
                    status_p = self._p("Registrado", 6, bold=True, color="#9a3412", align="CENTER")
                    table_style_commands_bg = ("BACKGROUND", (6, row_idx), (6, row_idx), colors.HexColor("#fef3c7"))
                
                data.append([
                    g_name,
                    self._p(apt_code, 7, align="CENTER"),
                    self._p(owner[:20] + "..." if len(owner) > 20 else owner, 7),
                    self._p(concept, 7),
                    self._money(amount),
                    self._p(method, 7),
                    status_p
                ])
                # We need to apply status background styling later
                span_commands.append(table_style_commands_bg)
                current_row_idx += 1
                
            if end_row > start_row:
                span_commands.append(("SPAN", (0, start_row), (0, end_row)))
                
        if len(data) == 1:
            data.append(["General", "-", "Sin pagos registrados", "-", self._money(0), "-", "-"])
            
        table_style_commands = [
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"), # Torre centered
            ("ALIGN", (1, 0), (1, -1), "CENTER"), # Depto centered
            ("ALIGN", (2, 0), (3, -1), "LEFT"),   # Owner & Concept left
            ("ALIGN", (4, 0), (4, -1), "RIGHT"),  # Monto right
            ("ALIGN", (5, 0), (5, -1), "LEFT"),   # Método left
            ("ALIGN", (6, 0), (6, -1), "CENTER"), # Estado centered
            ("GRID", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("ROWBACKGROUNDS", (1, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ] + span_commands
        
        detail_col_widths = [width * 0.12, width * 0.10, width * 0.22, width * 0.18, width * 0.14, width * 0.12, width * 0.12]
        story.append(Table(data, colWidths=detail_col_widths, style=TableStyle(table_style_commands)))
        story.append(Spacer(1, 0.18 * cm))
        
        # 4. Total bar
        icon_total = Drawing(20, 20)
        icon_total.add(Circle(10, 10, 9, fillColor=colors.white, strokeColor=None))
        icon_total.add(String(10, 6.5, "$", fontName="Helvetica-Bold", fontSize=10, fillColor=colors.HexColor("#0b3c7d"), textAnchor="middle"))
        
        val_total_box = Table([
            [icon_total, self._p("Total recaudado registrado en el período", 8, bold=True, color="#ffffff"), self._p(self._money(total_current), 9, bold=True, color="#ffffff", align="RIGHT")]
        ], colWidths=[24, width * 0.70, width * 0.30 - 24])
        val_total_box.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), _PDF_NAVY),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(val_total_box)
        story.append(Spacer(1, 0.22 * cm))
        
        # 5. Section: Comparativo respecto al mes anterior
        chart_icon = Drawing(16, 16)
        chart_icon.add(Circle(8, 8, 7.5, fillColor=colors.HexColor("#0b3c7d"), strokeColor=None))
        chart_icon.add(Rect(5, 4.5, 1.5, 4, fillColor=colors.white, strokeColor=None))
        chart_icon.add(Rect(7.25, 4.5, 1.5, 7, fillColor=colors.white, strokeColor=None))
        chart_icon.add(Rect(9.5, 4.5, 1.5, 5, fillColor=colors.white, strokeColor=None))
        
        comp_title_table = Table([[chart_icon, self._p("Comparativo respecto al mes anterior", 11, bold=True, color="#07316d", align="LEFT"), ""]], colWidths=[20, 240, width - 260])
        comp_title_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW", (2, 0), (2, 0), 1.2, _PDF_BLUE),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(comp_title_table)
        story.append(Spacer(1, 0.15 * cm))
        
        # Legend Row
        curr_period_name = self._period_name(current_period)
        prev_period_name = self._period_name(compare_period)
        legend_bar = self._build_legend(curr_period_name, prev_period_name)
        story.append(legend_bar)
        story.append(Spacer(1, 0.1 * cm))
        
        # Comparison Table
        max_val_money = max(total_current, total_compare, Decimal("1"))
        max_val_count = max(len(payments), len(compare_payments), 1)
        
        col_comp_w0 = width * 0.35
        col_comp_w1 = width * 0.45
        col_comp_w2 = width * 0.20
        
        comp_data = [
            [
                "", 
                "", 
                self._p("Valores", 7.5, bold=True, color="#475569", align="RIGHT")
            ],
            [
                self._p("Total recaudado (USD)", 7.5, bold=True, color="#1e293b"),
                self._draw_double_horizontal_bar(float(total_current), float(total_compare), float(max_val_money), col_comp_w1 - 12),
                self._stacked_monto(self._money(total_current), self._money(total_compare), col_comp_w2 - 12)
            ],
            [
                self._p("Cantidad de pagos", 7.5, bold=True, color="#1e293b"),
                self._draw_double_horizontal_bar(float(len(payments)), float(len(compare_payments)), float(max_val_count), col_comp_w1 - 12),
                self._stacked_monto(str(len(payments)), str(len(compare_payments)), col_comp_w2 - 12)
            ]
        ]
        
        comp_table = Table(comp_data, colWidths=[col_comp_w0, col_comp_w1, col_comp_w2])
        comp_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("LINEAFTER", (1, 0), (1, -1), 0.5, _PDF_BORDER),
            ("BOX", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]))
        story.append(comp_table)
        
        self._append_signature_grid(story, width=width, building=building, document_tag="REPORTE-PAGOS", signer_name="La Administración", signer_role="Periodo 2026-2027")
        footer = self._footer_callback(building, width)
        doc.build(story, onFirstPage=footer, onLaterPages=footer)
        return output.getvalue()

    async def expenses_pdf(
        self,
        period: Optional[str],
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        compare_period: Optional[str] = None,
    ) -> bytes:
        current_period = period or (start_date.strftime("%Y-%m") if start_date else date.today().strftime("%Y-%m"))
        current_start, current_end = self._get_period_dates(current_period)
        current_start = start_date or current_start
        current_end = end_date or current_end
        
        compare_period = compare_period or self._previous_period(current_period)
        compare_start, compare_end = self._get_period_dates(compare_period)
        
        expenses = await self._expenses(current_period, current_start, current_end)
        previous_expenses = await self._expenses(compare_period, compare_start, compare_end)
        
        total_current = sum(Decimal(str(e.get("amount", 0))) for e in expenses)
        total_compare = sum(Decimal(str(e.get("amount", 0))) for e in previous_expenses)
        
        by_category_current = self._build_breakdown(expenses, "category", "Sin categoría")
        previous_by_category = {row["label"]: row["amount"] for row in self._build_breakdown(previous_expenses, "category", "Sin categoría")}
        
        output = io.BytesIO()
        width = A4[0] - 2.2 * cm
        doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=1.1 * cm, rightMargin=1.1 * cm, topMargin=0.8 * cm, bottomMargin=0.7 * cm)
        story = []
        building = await get_default_building_config(getattr(self._payment_repo, "_conn", None))
        
        # Header
        story.extend(
            await self._three_column_report_header(
                "Reporte Detallado de Gastos",
                f"Rango: {self._date_label(period, start_date, end_date)} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                width,
                building=building,
                right_text=f"Total egresos: {self._money(total_current)}",
            )
        )
        
        # 1. 3 metric cards at the top
        icon_expenses = Drawing(24, 24)
        icon_expenses.add(Circle(12, 12, 11, fillColor=colors.HexColor("#e2e8f0"), strokeColor=colors.HexColor("#cbd5e1"), strokeWidth=0.5))
        icon_expenses.add(Rect(9, 6, 6, 12, fillColor=colors.white, strokeColor=colors.HexColor("#0b3c7d"), strokeWidth=1))
        icon_expenses.add(Line(11, 9, 13, 9, strokeColor=colors.HexColor("#0b3c7d"), strokeWidth=0.5))
        icon_expenses.add(Line(11, 11, 13, 11, strokeColor=colors.HexColor("#0b3c7d"), strokeWidth=0.5))
        icon_expenses.add(Line(11, 13, 13, 13, strokeColor=colors.HexColor("#0b3c7d"), strokeWidth=0.5))
        
        icon_categories = Drawing(24, 24)
        icon_categories.add(Circle(12, 12, 11, fillColor=colors.HexColor("#e2e8f0"), strokeColor=colors.HexColor("#cbd5e1"), strokeWidth=0.5))
        icon_categories.add(Rect(9, 9, 2.5, 2.5, fillColor=colors.HexColor("#0b3c7d"), strokeColor=None))
        icon_categories.add(Rect(12.5, 9, 2.5, 2.5, fillColor=colors.HexColor("#0b3c7d"), strokeColor=None))
        icon_categories.add(Rect(9, 12.5, 2.5, 2.5, fillColor=colors.HexColor("#0b3c7d"), strokeColor=None))
        icon_categories.add(Rect(12.5, 12.5, 2.5, 2.5, fillColor=colors.HexColor("#0b3c7d"), strokeColor=None))
        
        col_w_card = (width / 3) - 5
        cards = [
            self._build_income_metric_card("Gastos", str(len(expenses)), "G", col_w_card),
            self._build_income_metric_card("Total", self._money(total_current), "$", col_w_card),
            self._build_income_metric_card("Categorías", str(len(by_category_current)), "C", col_w_card),
        ]
        # Override Card 1 & 3 icons with vector shapes
        cards[0] = Table([
            [self._p("GASTOS", 6.5, bold=True, color="#ffffff", align="CENTER")],
            [Table([
                [icon_expenses, self._p(str(len(expenses)), 11, bold=True, color="#0b3c7d", align="LEFT")]
            ], colWidths=[28, col_w_card - 38], style=[
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ])]
        ], colWidths=[col_w_card])
        cards[0].setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        
        cards[2] = Table([
            [self._p("CATEGORÍAS", 6.5, bold=True, color="#ffffff", align="CENTER")],
            [Table([
                [icon_categories, self._p(str(len(by_category_current)), 11, bold=True, color="#0b3c7d", align="LEFT")]
            ], colWidths=[28, col_w_card - 38], style=[
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ])]
        ], colWidths=[col_w_card])
        cards[2].setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        
        cards_table = Table([cards], colWidths=[width / 3] * 3)
        cards_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(cards_table)
        story.append(Spacer(1, 0.25 * cm))
        
        # 2. Custom Title with List Icon
        list_icon = Drawing(16, 16)
        list_icon.add(Circle(8, 8, 7.5, fillColor=colors.HexColor("#0b3c7d"), strokeColor=None))
        list_icon.add(Line(5, 10, 11, 10, strokeColor=colors.white, strokeWidth=1))
        list_icon.add(Line(5, 8, 11, 8, strokeColor=colors.white, strokeWidth=1))
        list_icon.add(Line(5, 6, 11, 6, strokeColor=colors.white, strokeWidth=1))
        
        title_table = Table([[list_icon, self._p("Detalle de gastos", 11, bold=True, color="#07316d", align="LEFT"), ""]], colWidths=[20, 200, width - 220])
        title_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW", (2, 0), (2, 0), 1.2, _PDF_BLUE),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(title_table)
        story.append(Spacer(1, 0.15 * cm))
        
        # 3. Detalle de gastos table
        data = [[
            self._table_p("Fecha", 7.2, bold=True, color="#ffffff", align="CENTER"),
            self._table_p("Proveedor", 7.2, bold=True, color="#ffffff"),
            self._table_p("Categoría", 7.2, bold=True, color="#ffffff"),
            self._table_p("Concepto", 7.2, bold=True, color="#ffffff"),
            self._table_p("Monto", 7.2, bold=True, color="#ffffff", align="RIGHT"),
            self._table_p("Comprobante", 7.2, bold=True, color="#ffffff"),
        ]]
        for e in expenses:
            comp_str = e.get("receipt_file_name") or "Sin comprobante"
            
            dt = e.get("date")
            date_str = dt.strftime("%Y-%m-%d") if hasattr(dt, "strftime") else str(dt or "")
            
            data.append([
                self._table_p(date_str, 7, align="CENTER"),
                self._table_p(e.get("provider") or "", 7),
                self._table_p(e.get("category") or "Sin categoría", 7),
                self._table_p(e.get("concept") or "", 7),
                self._table_p(self._money(e.get("amount", 0)), 7, align="RIGHT"),
                self._table_p(comp_str, 7),
            ])
        if len(data) == 1:
            data.append([
                self._table_p("-", 7, align="CENTER"),
                self._table_p("-", 7),
                self._table_p("-", 7),
                self._table_p("Sin gastos registrados", 7),
                self._table_p(self._money(0), 7, align="RIGHT"),
                self._table_p("-", 7),
            ])
            
        table_style_commands = [
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (1, 0), (2, -1), "LEFT"),
            ("ALIGN", (3, 0), (3, -1), "LEFT"),
            ("ALIGN", (4, 0), (4, -1), "RIGHT"),
            ("ALIGN", (5, 0), (5, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("WORDWRAP", (0, 0), (-1, -1), "CJK"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]
        
        detail_col_widths = [
            width * 0.13,
            width * 0.21,
            width * 0.16,
            width * 0.28,
            width * 0.09,
            width * 0.13,
        ]
        story.append(Table(data, colWidths=detail_col_widths, repeatRows=1, style=TableStyle(table_style_commands)))
        story.append(Spacer(1, 0.18 * cm))
        
        # 4. Total bar
        icon_total = Drawing(20, 20)
        icon_total.add(Circle(10, 10, 9, fillColor=colors.white, strokeColor=None))
        icon_total.add(String(10, 6.5, "$", fontName="Helvetica-Bold", fontSize=10, fillColor=colors.HexColor("#0b3c7d"), textAnchor="middle"))
        
        val_total_box = Table([
            [icon_total, self._p("Egresos totales registrados en el período", 8, bold=True, color="#ffffff"), self._p(self._money(total_current), 9, bold=True, color="#ffffff", align="RIGHT")]
        ], colWidths=[24, width * 0.70, width * 0.30 - 24])
        val_total_box.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), _PDF_NAVY),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(val_total_box)
        story.append(Spacer(1, 0.22 * cm))
        
        # 5. Section: Comparativo respecto al mes anterior
        chart_icon = Drawing(16, 16)
        chart_icon.add(Circle(8, 8, 7.5, fillColor=colors.HexColor("#0b3c7d"), strokeColor=None))
        chart_icon.add(Rect(5, 4.5, 1.5, 4, fillColor=colors.white, strokeColor=None))
        chart_icon.add(Rect(7.25, 4.5, 1.5, 7, fillColor=colors.white, strokeColor=None))
        chart_icon.add(Rect(9.5, 4.5, 1.5, 5, fillColor=colors.white, strokeColor=None))
        
        comp_title_table = Table([[chart_icon, self._p("Comparativo respecto al mes anterior", 11, bold=True, color="#07316d", align="LEFT"), ""]], colWidths=[20, 240, width - 260])
        comp_title_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW", (2, 0), (2, 0), 1.2, _PDF_BLUE),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(comp_title_table)
        story.append(Spacer(1, 0.15 * cm))
        
        # Legend Row
        curr_period_name = self._period_name(current_period)
        prev_period_name = self._period_name(compare_period)
        legend_bar = self._build_legend(curr_period_name, prev_period_name)
        story.append(legend_bar)
        story.append(Spacer(1, 0.1 * cm))
        
        # Comparison Table
        max_val = max(total_current, total_compare, Decimal("1"))
        
        col_comp_w0 = width * 0.35
        col_comp_w1 = width * 0.45
        col_comp_w2 = width * 0.20
        
        comp_data = [
            [
                "", 
                "", 
                self._p("Valores (USD)", 7.5, bold=True, color="#475569", align="RIGHT")
            ],
            [
                self._p("Egresos totales registrados en el período", 7.5, bold=True, color="#1e293b"),
                self._draw_double_horizontal_bar(float(total_current), float(total_compare), float(max_val), col_comp_w1 - 12),
                self._stacked_monto(self._money(total_current), self._money(total_compare), col_comp_w2 - 12)
            ]
        ]
        for c in by_category_current[:4]:
            cat_label = c["label"]
            cat_curr = c["amount"]
            cat_prev = previous_by_category.get(cat_label, Decimal("0"))
            
            comp_data.append([
                self._p(cat_label, 7.5, bold=True, color="#1e293b"),
                self._draw_double_horizontal_bar(float(cat_curr), float(cat_prev), float(max_val), col_comp_w1 - 12),
                self._stacked_monto(self._money(cat_curr), self._money(cat_prev), col_comp_w2 - 12)
            ])
            
        comp_table = Table(comp_data, colWidths=[col_comp_w0, col_comp_w1, col_comp_w2])
        comp_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("LINEAFTER", (1, 0), (1, -1), 0.5, _PDF_BORDER),
            ("BOX", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]))
        story.append(comp_table)
        
        self._append_signature_grid(story, width=width, building=building, document_tag="REPORTE-GASTOS", signer_name="La Administración", signer_role="Periodo 2026-2027")
        footer = self._footer_callback(building, width)
        doc.build(story, onFirstPage=footer, onLaterPages=footer)
        return output.getvalue()

    async def delinquency_pdf(self) -> bytes:
        owners = await self._delinquency.list_owners()
        building = await get_default_building_config(getattr(self._payment_repo, "_conn", None))

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=letter,
            leftMargin=0.45 * inch,
            rightMargin=0.45 * inch,
            topMargin=0.45 * inch,
            bottomMargin=0.75 * inch,
        )
        story = []
        story.extend(
            await self._three_column_report_header(
                "Reporte de Morosidad",
                f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                width=doc.width,
                building=building,
                right_text=f"Propietarios: {len(owners)}",
            )
        )

        data = [[
            self._table_p("Propietario", 7, bold=True, color="#ffffff", align="CENTER"),
            self._table_p("Email", 7, bold=True, color="#ffffff", align="CENTER"),
            self._table_p("Documento", 7, bold=True, color="#ffffff", align="CENTER"),
            self._table_p("Deuda Total", 7, bold=True, color="#ffffff", align="RIGHT"),
            self._table_p("Períodos Vencidos", 7, bold=True, color="#ffffff", align="CENTER"),
            self._table_p("Estado", 7, bold=True, color="#ffffff", align="CENTER"),
        ]]
        for o in owners:
            data.append([
                self._table_p(o["owner_name"], 6.4, color="#1e293b"),
                self._table_p(o.get("email") or "", 6.2, color="#1e293b"),
                self._table_p(o["document_id"], 6.4, color="#1e293b", align="CENTER"),
                self._table_p(self._money(o["deuda_total"]), 6.4, bold=True, color="#1e293b", align="RIGHT"),
                self._table_p(str(o["periodos_vencidos"]), 6.4, color="#1e293b", align="CENTER"),
                self._table_p(self._delinquency_status_label(o["status"]), 6.4, bold=True, color="#1e293b", align="CENTER"),
            ])

        table = Table(
            data,
            colWidths=[
                doc.width * 0.24,
                doc.width * 0.27,
                doc.width * 0.13,
                doc.width * 0.12,
                doc.width * 0.13,
                doc.width * 0.11,
            ],
            repeatRows=1,
        )
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("ALIGN", (0, 1), (1, -1), "LEFT"),
            ("ALIGN", (2, 1), (-1, -1), "CENTER"),
            ("ALIGN", (3, 1), (3, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.45, _PDF_BORDER),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(table)
        story.append(Spacer(1, 0.22 * inch))
        story.append(self._p("Reporte generado automáticamente", 8, color="#334155"))
        self._append_signature_grid(story, width=doc.width, building=building, document_tag="REPORTE-MOROSIDAD", signer_name="Administración", signer_role="Administrador del edificio")
        footer = self._footer_callback(building, doc.width)

        doc.build(story, onFirstPage=footer, onLaterPages=footer)
        return output.getvalue()

    # ─── EXCEL REPORTS ───────────────────────────────────────────────────────

    async def income_excel(
        self,
        period: Optional[str],
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        payments = await self._income_entries(period, start_date, end_date)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Ingresos"
        
        # Headers
        headers = ["Fecha", "Origen", "Concepto", "Propietario", "Departamento", "Período", "Monto", "Método", "Estado"]
        header_fill = PatternFill(start_color="123C7A", end_color="123C7A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row, p in enumerate(payments, 2):
            ws.cell(row=row, column=1).value = str(p.get("income_date") or p.get("paid_at") or "")
            ws.cell(row=row, column=2).value = p.get("income_label", "")
            ws.cell(row=row, column=3).value = p.get("income_concept", "")
            ws.cell(row=row, column=4).value = p.get("owner_name", "")
            ws.cell(row=row, column=5).value = p.get("apartment_code", "")
            ws.cell(row=row, column=6).value = p.get("period", "")
            ws.cell(row=row, column=7).value = float(p.get("amount", 0))
            ws.cell(row=row, column=8).value = p.get("method", "")
            ws.cell(row=row, column=9).value = p.get("status", "")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def balance_excel(
        self,
        period: Optional[str],
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        payments = await self._income_entries(period, start_date, end_date)
        expenses = await self._expenses(period, start_date, end_date)

        total_income = sum(Decimal(str(p.get("amount", 0))) for p in payments)
        total_expenses = sum(Decimal(str(e.get("amount", 0))) for e in expenses)
        balance = total_income - total_expenses
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen"
        headers = ["Indicador", "Valor", "Detalle"]
        self._style_excel_header(ws, headers)
        rows = [
            ["Ingresos confirmados", float(total_income), f"{len(payments)} movimientos"],
            ["Gastos registrados", float(total_expenses), f"{len(expenses)} gastos"],
            ["Diferencia neta", float(balance), "Ingresos menos gastos"],
        ]
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx).value = value
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 28

        income_ws = wb.create_sheet("Ingresos")
        self._style_excel_header(income_ws, ["Fecha", "Origen", "Concepto", "Propietario", "Departamento", "Período", "Monto", "Método", "Estado", "Referencia"])
        for row, p in enumerate(payments, 2):
            values = [
                str(p.get("income_date") or p.get("paid_at") or ""),
                p.get("income_label") or "",
                p.get("income_concept") or "",
                p.get("owner_name") or "",
                p.get("apartment_code") or "",
                p.get("period") or "",
                float(p.get("amount", 0)),
                p.get("method") or "",
                p.get("status") or "",
                p.get("reference") or "",
            ]
            for col, value in enumerate(values, 1):
                income_ws.cell(row=row, column=col).value = value
        for col in "ABCDEFGHIJ":
            income_ws.column_dimensions[col].width = 16

        expense_ws = wb.create_sheet("Gastos")
        self._style_excel_header(expense_ws, ["Fecha", "Proveedor", "Categoría", "Concepto", "Monto", "Comprobante"])
        for row, e in enumerate(expenses, 2):
            values = [str(e.get("date") or ""), e.get("provider") or "", e.get("category") or "Sin categoría", e.get("concept") or "", float(e.get("amount", 0)), e.get("receipt_file_name") or ""]
            for col, value in enumerate(values, 1):
                expense_ws.cell(row=row, column=col).value = value
        for col in "ABCDEF":
            expense_ws.column_dimensions[col].width = 18
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def payments_excel(
        self,
        period: Optional[str],
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        status: Optional[str] = None,
    ) -> bytes:
        payments = await self._payments_report_rows(period, start_date, end_date, status)
        wb = Workbook()
        ws = wb.active
        ws.title = "Pagos"
        headers = ["Fecha", "Propietario", "Departamento", "Período", "Monto", "Método", "Estado", "Referencia"]
        self._style_excel_header(ws, headers)
        for row, p in enumerate(payments, 2):
            values = [str(p.get("paid_at") or ""), p.get("owner_name") or "", p.get("apartment_code") or "", p.get("period") or "", float(p.get("amount", 0)), p.get("method") or "", p.get("status") or "", p.get("reference") or ""]
            for col, value in enumerate(values, 1):
                ws.cell(row=row, column=col).value = value
        for col in "ABCDEFGH":
            ws.column_dimensions[col].width = 16
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def expenses_excel(
        self,
        period: Optional[str],
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        expenses = await self._expenses(period, start_date, end_date)
        wb = Workbook()
        ws = wb.active
        ws.title = "Gastos"
        headers = ["Fecha", "Proveedor", "Categoría", "Concepto", "Monto", "Comprobante"]
        self._style_excel_header(ws, headers)
        for row, e in enumerate(expenses, 2):
            values = [str(e.get("date") or ""), e.get("provider") or "", e.get("category") or "Sin categoría", e.get("concept") or "", float(e.get("amount", 0)), e.get("receipt_file_name") or ""]
            for col, value in enumerate(values, 1):
                ws.cell(row=row, column=col).value = value
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 34
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 24
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def _fees_report_rows(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> list[dict]:
        conditions: list[str] = []
        params: list = []
        idx = 1
        if start_date:
            conditions.append(f"af.period >= ${idx}")
            params.append(start_date.strftime("%Y-%m"))
            idx += 1
        if end_date:
            conditions.append(f"af.period <= ${idx}")
            params.append(end_date.strftime("%Y-%m"))
        where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
        rows = await self._payment_repo._conn.fetch(
            f"""
            SELECT
                af.period,
                a.code AS apartment_code,
                a.tower AS apartment_tower,
                COALESCE(o.full_name, 'Sin propietario') AS owner_name,
                af.amount,
                COALESCE(p.paid_amount, 0) AS paid_amount,
                af.amount - COALESCE(p.paid_amount, 0) AS pending_amount,
                CASE WHEN COALESCE(p.paid_amount, 0) >= af.amount THEN 'PAGADA' ELSE 'PENDIENTE' END AS status
            FROM apartment_fees af
            JOIN apartments a ON af.apartment_id = a.id
            LEFT JOIN owner_apartments oa ON oa.apartment_id = a.id
            LEFT JOIN owners o ON o.id = oa.owner_id
            LEFT JOIN (
                SELECT apartment_id, period, SUM(amount) AS paid_amount
                FROM payments
                WHERE status IN ('REGISTRADO', 'APROBADO') AND fine_id IS NULL
                GROUP BY apartment_id, period
            ) p ON p.apartment_id = af.apartment_id AND p.period = af.period
            {where}
            ORDER BY af.period DESC, a.code
            """,
            *params,
        )
        return [dict(row) for row in rows]

    async def _fines_report_rows(
        self,
        period: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        status: Optional[str] = None,
        reason: Optional[str] = None,
        search: Optional[str] = None,
    ) -> list[dict]:
        conditions: list[str] = []
        params: list = []
        idx = 1
        if period:
            conditions.append(f"f.period = ${idx}")
            params.append(period)
            idx += 1
        if status:
            conditions.append(f"f.status = ${idx}")
            params.append(status)
            idx += 1
        if reason:
            conditions.append(f"f.reason = ${idx}")
            params.append(reason)
            idx += 1
        if search:
            conditions.append(
                f"(LOWER(COALESCE(f.reason, '')) LIKE ${idx} OR LOWER(a.code) LIKE ${idx} OR LOWER(o.full_name) LIKE ${idx})"
            )
            params.append(f"%{search.lower()}%")
            idx += 1
        if start_date:
            conditions.append(f"f.issued_at >= ${idx}")
            params.append(start_date)
            idx += 1
        if end_date:
            conditions.append(f"f.issued_at <= ${idx}")
            params.append(end_date)
        where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
        rows = await self._payment_repo._conn.fetch(
            f"""
            SELECT f.*, a.code AS apartment_code, o.full_name AS owner_name
            FROM fines f
            JOIN apartments a ON f.apartment_id = a.id
            JOIN owners o ON f.owner_id = o.id
            {where}
            ORDER BY f.issued_at DESC, f.created_at DESC
            """,
            *params,
        )
        return [dict(row) for row in rows]

    async def _owners_report_rows(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        status: Optional[str] = None,
    ) -> list[dict]:
        conditions: list[str] = []
        params: list = []
        idx = 1
        if status:
            conditions.append(f"o.status = ${idx}")
            params.append(status)
            idx += 1
        if start_date:
            conditions.append(f"o.created_at::date >= ${idx}")
            params.append(start_date)
            idx += 1
        if end_date:
            conditions.append(f"o.created_at::date <= ${idx}")
            params.append(end_date)
        where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
        rows = await self._payment_repo._conn.fetch(
            f"""
            SELECT
                o.id,
                o.full_name,
                o.document_id,
                o.email,
                o.phone,
                o.status,
                o.created_at,
                COUNT(DISTINCT oa.apartment_id)::int AS units_count,
                STRING_AGG(DISTINCT a.code, ', ' ORDER BY a.code) AS units
            FROM owners o
            LEFT JOIN owner_apartments oa ON oa.owner_id = o.id
            LEFT JOIN apartments a ON a.id = oa.apartment_id
            {where}
            GROUP BY o.id
            ORDER BY o.full_name
            """,
            *params,
        )
        return [dict(row) for row in rows]

    async def _buildings_report_rows(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> list[dict]:
        conditions: list[str] = []
        params: list = []
        idx = 1
        if start_date:
            conditions.append(f"b.created_at::date >= ${idx}")
            params.append(start_date)
            idx += 1
        if end_date:
            conditions.append(f"b.created_at::date <= ${idx}")
            params.append(end_date)
        where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
        rows = await self._payment_repo._conn.fetch(
            f"""
            SELECT
                b.*,
                COUNT(DISTINCT a.id)::int AS apartments_count
            FROM buildings b
            LEFT JOIN apartments a ON a.building_id = b.id
            {where}
            GROUP BY b.id
            ORDER BY b.created_at ASC
            """,
            *params,
        )
        return [dict(row) for row in rows]

    async def fees_pdf(
        self,
        period: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        compare_period: Optional[str] = None,
    ) -> bytes:
        current_period = period or (start_date.strftime("%Y-%m") if start_date else date.today().strftime("%Y-%m"))
        current_start, current_end = self._get_period_dates(current_period)
        current_start = start_date or current_start
        current_end = end_date or current_end
        
        compare_period = compare_period or self._previous_period(current_period)
        compare_start, compare_end = self._get_period_dates(compare_period)
        
        rows = await self._fees_report_rows(current_start, current_end)
        compare_rows = await self._fees_report_rows(compare_start, compare_end)
        
        total_current = sum(Decimal(str(row.get("amount", 0))) for row in rows)
        collected_current = sum(Decimal(str(row.get("paid_amount", 0))) for row in rows)
        pending_current = total_current - collected_current
        
        total_compare = sum(Decimal(str(row.get("amount", 0))) for row in compare_rows)
        collected_compare = sum(Decimal(str(row.get("paid_amount", 0))) for row in compare_rows)
        pending_compare = total_compare - collected_compare
        
        building = await get_default_building_config(getattr(self._payment_repo, "_conn", None))
        output = io.BytesIO()
        width = A4[0] - 2.2 * cm
        doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=1.1 * cm, rightMargin=1.1 * cm, topMargin=0.8 * cm, bottomMargin=0.7 * cm)
        story = []
        
        # Header
        story.extend(
            await self._three_column_report_header(
                "Reporte Detallado de Cuotas",
                f"Rango: {self._date_label(period, start_date, end_date)} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                width,
                building=building,
                right_text=f"Total emitido: {self._money(total_current)}",
            )
        )
        
        # 1. Metric cards
        icon_cuotas = Drawing(24, 24)
        icon_cuotas.add(Circle(12, 12, 11, fillColor=colors.HexColor("#e2e8f0"), strokeColor=colors.HexColor("#cbd5e1"), strokeWidth=0.5))
        icon_cuotas.add(Rect(9, 6, 6, 12, fillColor=colors.white, strokeColor=colors.HexColor("#0b3c7d"), strokeWidth=1))
        icon_cuotas.add(Line(11, 9, 13, 9, strokeColor=colors.HexColor("#0b3c7d"), strokeWidth=0.5))
        icon_cuotas.add(Line(11, 12, 13, 12, strokeColor=colors.HexColor("#0b3c7d"), strokeWidth=0.5))
        
        icon_emitido = Drawing(24, 24)
        icon_emitido.add(Circle(12, 12, 11, fillColor=colors.HexColor("#e2e8f0"), strokeColor=colors.HexColor("#cbd5e1"), strokeWidth=0.5))
        icon_emitido.add(String(12, 8, "$", fontName="Helvetica-Bold", fontSize=11, fillColor=colors.HexColor("#0b3c7d"), textAnchor="middle"))
        
        icon_recaudado = Drawing(24, 24)
        icon_recaudado.add(Circle(12, 12, 11, fillColor=colors.HexColor("#e2e8f0"), strokeColor=colors.HexColor("#cbd5e1"), strokeWidth=0.5))
        icon_recaudado.add(Line(8, 12, 11, 9, strokeColor=colors.HexColor("#166534"), strokeWidth=1.5))
        icon_recaudado.add(Line(11, 9, 16, 15, strokeColor=colors.HexColor("#166534"), strokeWidth=1.5))
        
        icon_pendiente = Drawing(24, 24)
        icon_pendiente.add(Circle(12, 12, 11, fillColor=colors.HexColor("#e2e8f0"), strokeColor=colors.HexColor("#cbd5e1"), strokeWidth=0.5))
        icon_pendiente.add(Circle(12, 12, 5, fillColor=None, strokeColor=colors.HexColor("#b91c1c"), strokeWidth=1))
        icon_pendiente.add(Line(12, 12, 12, 15, strokeColor=colors.HexColor("#b91c1c"), strokeWidth=1))
        icon_pendiente.add(Line(12, 12, 14, 12, strokeColor=colors.HexColor("#b91c1c"), strokeWidth=1))
        
        col_w_card = (width / 4) - 5
        
        def make_card(title: str, val_str: str, icon_flowable):
            fs = 10
            if len(val_str) > 10:
                fs = 8.5
            if len(val_str) > 13:
                fs = 7.5
            t = Table([
                [self._p(title.upper(), 6.5, bold=True, color="#ffffff", align="CENTER")],
                [Table([
                    [icon_flowable, self._p(val_str, fs, bold=True, color="#0b3c7d", align="LEFT")]
                ], colWidths=[26, col_w_card - 36], style=[
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ])]
            ], colWidths=[col_w_card])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]))
            return t
            
        cards = [
            make_card("Cuotas", str(len(rows)), icon_cuotas),
            make_card("Emitido", self._money(total_current), icon_emitido),
            make_card("Recaudado", self._money(collected_current), icon_recaudado),
            make_card("Pendiente", self._money(pending_current), icon_pendiente),
        ]
        cards_table = Table([cards], colWidths=[width / 4] * 4)
        cards_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(cards_table)
        story.append(Spacer(1, 0.25 * cm))
        
        # 2. Custom Title with List Icon
        list_icon = Drawing(16, 16)
        list_icon.add(Circle(8, 8, 7.5, fillColor=colors.HexColor("#0b3c7d"), strokeColor=None))
        list_icon.add(Line(5, 10, 11, 10, strokeColor=colors.white, strokeWidth=1))
        list_icon.add(Line(5, 8, 11, 8, strokeColor=colors.white, strokeWidth=1))
        list_icon.add(Line(5, 6, 11, 6, strokeColor=colors.white, strokeWidth=1))
        
        title_table = Table([[list_icon, self._p("Detalle de cuotas", 11, bold=True, color="#07316d", align="LEFT"), ""]], colWidths=[20, 200, width - 220])
        title_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW", (2, 0), (2, 0), 1.2, _PDF_BLUE),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(title_table)
        story.append(Spacer(1, 0.15 * cm))
        
        # 3. Detalle de cuotas table
        data = [[
            self._table_p("Período", 7.2, bold=True, color="#ffffff", align="CENTER"),
            self._table_p("Depto", 7.2, bold=True, color="#ffffff", align="CENTER"),
            self._table_p("Propietario", 7.2, bold=True, color="#ffffff"),
            self._table_p("Emitido", 7.2, bold=True, color="#ffffff", align="RIGHT"),
            self._table_p("Pagado", 7.2, bold=True, color="#ffffff", align="RIGHT"),
            self._table_p("Pendiente", 7.2, bold=True, color="#ffffff", align="RIGHT"),
            self._table_p("Estado", 7.2, bold=True, color="#ffffff", align="CENTER"),
        ]]
        span_commands = []
        for r in rows:
            row_idx = len(data)
            status_str = r.get("status") or "PENDIENTE"
            
            if status_str == "PAGADA":
                status_p = self._p("Pagada", 6, bold=True, color="#166534", align="CENTER")
                table_style_commands_bg = ("BACKGROUND", (6, row_idx), (6, row_idx), colors.HexColor("#dcfce7"))
            else:
                status_p = self._p("Pendiente", 6, bold=True, color="#9a3412", align="CENTER")
                table_style_commands_bg = ("BACKGROUND", (6, row_idx), (6, row_idx), colors.HexColor("#fef3c7"))
                
            data.append([
                self._table_p(r.get("period", ""), 7, align="CENTER"),
                self._table_p(r.get("apartment_code", ""), 7, align="CENTER"),
                self._table_p(r.get("owner_name") or "", 7),
                self._table_p(self._money(r.get("amount")), 7, align="RIGHT"),
                self._table_p(self._money(r.get("paid_amount")), 7, align="RIGHT"),
                self._table_p(self._money(r.get("pending_amount")), 7, align="RIGHT"),
                status_p
            ])
            span_commands.append(table_style_commands_bg)
            
        if len(data) == 1:
            data.append([
                self._table_p("-", 7, align="CENTER"),
                self._table_p("-", 7, align="CENTER"),
                self._table_p("Sin cuotas emitidas", 7),
                self._table_p(self._money(0), 7, align="RIGHT"),
                self._table_p(self._money(0), 7, align="RIGHT"),
                self._table_p(self._money(0), 7, align="RIGHT"),
                self._table_p("-", 7, align="CENTER"),
            ])
            
        table_style_commands = [
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (1, -1), "CENTER"), # Period & Depto centered
            ("ALIGN", (2, 0), (2, -1), "LEFT"),   # Owner left
            ("ALIGN", (3, 0), (5, -1), "RIGHT"),  # Amounts right
            ("ALIGN", (6, 0), (6, -1), "CENTER"), # Estado centered
            ("GRID", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("WORDWRAP", (0, 0), (-1, -1), "CJK"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ] + span_commands
        
        detail_col_widths = [width * 0.12, width * 0.10, width * 0.28, width * 0.12, width * 0.12, width * 0.12, width * 0.14]
        story.append(Table(data, colWidths=detail_col_widths, repeatRows=1, style=TableStyle(table_style_commands)))
        story.append(Spacer(1, 0.18 * cm))
        
        # 4. Total bar
        icon_total = Drawing(20, 20)
        icon_total.add(Circle(10, 10, 9, fillColor=colors.white, strokeColor=None))
        icon_total.add(String(10, 6.5, "$", fontName="Helvetica-Bold", fontSize=10, fillColor=colors.HexColor("#0b3c7d"), textAnchor="middle"))
        
        val_total_box = Table([
            [icon_total, self._p("Resumen de cuotas emitidas en el período", 8, bold=True, color="#ffffff"), self._p(self._money(total_current), 9, bold=True, color="#ffffff", align="RIGHT")]
        ], colWidths=[24, width * 0.70, width * 0.30 - 24])
        val_total_box.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), _PDF_NAVY),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(val_total_box)
        story.append(Spacer(1, 0.22 * cm))
        
        # 5. Section: Comparativo respecto al mes anterior
        chart_icon = Drawing(16, 16)
        chart_icon.add(Circle(8, 8, 7.5, fillColor=colors.HexColor("#0b3c7d"), strokeColor=None))
        chart_icon.add(Rect(5, 4.5, 1.5, 4, fillColor=colors.white, strokeColor=None))
        chart_icon.add(Rect(7.25, 4.5, 1.5, 7, fillColor=colors.white, strokeColor=None))
        chart_icon.add(Rect(9.5, 4.5, 1.5, 5, fillColor=colors.white, strokeColor=None))
        
        comp_title_table = Table([[chart_icon, self._p("Comparativo respecto al mes anterior", 11, bold=True, color="#07316d", align="LEFT"), ""]], colWidths=[20, 240, width - 260])
        comp_title_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW", (2, 0), (2, 0), 1.2, _PDF_BLUE),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(comp_title_table)
        story.append(Spacer(1, 0.15 * cm))
        
        # Legend Row
        curr_period_name = self._period_name(current_period)
        prev_period_name = self._period_name(compare_period)
        legend_bar = self._build_legend(curr_period_name, prev_period_name)
        story.append(legend_bar)
        story.append(Spacer(1, 0.1 * cm))
        
        # Comparison Table
        max_val_money = max(total_current, total_compare, collected_current, collected_compare, pending_current, pending_compare, Decimal("1"))
        
        col_comp_w0 = width * 0.35
        col_comp_w1 = width * 0.45
        col_comp_w2 = width * 0.20
        
        comp_data = [
            [
                "", 
                "", 
                self._p("Valores (USD)", 7.5, bold=True, color="#475569", align="RIGHT")
            ],
            [
                self._p("Cuotas emitidas", 7.5, bold=True, color="#1e293b"),
                self._draw_double_horizontal_bar(float(total_current), float(total_compare), float(max_val_money), col_comp_w1 - 12),
                self._stacked_monto(self._money(total_current), self._money(total_compare), col_comp_w2 - 12)
            ],
            [
                self._p("Cuotas recaudadas", 7.5, bold=True, color="#1e293b"),
                self._draw_double_horizontal_bar(float(collected_current), float(collected_compare), float(max_val_money), col_comp_w1 - 12),
                self._stacked_monto(self._money(collected_current), self._money(collected_compare), col_comp_w2 - 12)
            ],
            [
                self._p("Saldo pendiente", 7.5, bold=True, color="#1e293b"),
                self._draw_double_horizontal_bar(float(pending_current), float(pending_compare), float(max_val_money), col_comp_w1 - 12),
                self._stacked_monto(self._money(pending_current), self._money(pending_compare), col_comp_w2 - 12)
            ]
        ]
        
        comp_table = Table(comp_data, colWidths=[col_comp_w0, col_comp_w1, col_comp_w2])
        comp_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("LINEAFTER", (1, 0), (1, -1), 0.5, _PDF_BORDER),
            ("BOX", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]))
        story.append(comp_table)
        
        self._append_signature_grid(story, width=width, building=building, document_tag="REPORTE-CUOTAS", signer_name="La Administración", signer_role="Periodo 2026-2027")
        footer = self._footer_callback(building, width)
        doc.build(story, onFirstPage=footer, onLaterPages=footer)
        return output.getvalue()

    async def fees_excel(self, start_date: Optional[date] = None, end_date: Optional[date] = None) -> bytes:
        rows = await self._fees_report_rows(start_date, end_date)
        wb = Workbook()
        ws = wb.active
        ws.title = "Cuotas"
        headers = ["Período", "Departamento", "Propietario", "Emitido", "Pagado", "Pendiente", "Estado"]
        self._style_excel_header(ws, headers)
        for row_idx, row in enumerate(rows, 2):
            values = [row.get("period", ""), row.get("apartment_code", ""), row.get("owner_name", ""), float(row.get("amount", 0)), float(row.get("paid_amount", 0)), float(row.get("pending_amount", 0)), row.get("status", "")]
            for col, value in enumerate(values, 1):
                ws.cell(row=row_idx, column=col).value = value
        for col in "ABCDEFG":
            ws.column_dimensions[col].width = 18
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def fines_pdf(self, period: Optional[str] = None, start_date: Optional[date] = None, end_date: Optional[date] = None, status: Optional[str] = None, reason: Optional[str] = None, search: Optional[str] = None) -> bytes:
        rows = await self._fines_report_rows(period, start_date, end_date, status, reason, search)
        total = sum(Decimal(str(row.get("amount", 0))) for row in rows)
        building = await get_default_building_config(getattr(self._payment_repo, "_conn", None))
        output = io.BytesIO()
        width = A4[0] - 2.2 * cm
        doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=1.1 * cm, rightMargin=1.1 * cm, topMargin=0.8 * cm, bottomMargin=0.7 * cm)
        story = []
        story.extend(
            await self._three_column_report_header(
                "Reporte Detallado de Multas",
                f"Rango: {self._date_label(period, start_date, end_date)} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                width,
                building=building,
                right_text=f"Estado: {status or 'Todos'}",
            )
        )

        summary = Table(
            [[
                self._table_p("Multas", 7.2, bold=True, color="#ffffff", align="CENTER"),
                self._table_p("Monto total", 7.2, bold=True, color="#ffffff", align="RIGHT"),
            ], [
                self._table_p(str(len(rows)), 7, align="CENTER"),
                self._table_p(self._money(total), 7, align="RIGHT"),
            ]],
            colWidths=[width * 0.50, width * 0.50],
        )
        summary.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]))
        story.extend([summary, Spacer(1, 0.25 * cm)])

        list_icon = Drawing(16, 16)
        list_icon.add(Circle(8, 8, 7.5, fillColor=colors.HexColor("#0b3c7d"), strokeColor=None))
        list_icon.add(Line(5, 10, 11, 10, strokeColor=colors.white, strokeWidth=1))
        list_icon.add(Line(5, 8, 11, 8, strokeColor=colors.white, strokeWidth=1))
        list_icon.add(Line(5, 6, 11, 6, strokeColor=colors.white, strokeWidth=1))
        title_table = Table([[list_icon, self._p("Detalle de multas", 11, bold=True, color="#07316d", align="LEFT"), ""]], colWidths=[20, 200, width - 220])
        title_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW", (2, 0), (2, 0), 1.2, _PDF_BLUE),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.extend([title_table, Spacer(1, 0.15 * cm)])

        data = [[
            self._table_p("Fecha", 7.2, bold=True, color="#ffffff", align="CENTER"),
            self._table_p("Depto", 7.2, bold=True, color="#ffffff", align="CENTER"),
            self._table_p("Propietario", 7.2, bold=True, color="#ffffff"),
            self._table_p("Período", 7.2, bold=True, color="#ffffff", align="CENTER"),
            self._table_p("Motivo", 7.2, bold=True, color="#ffffff"),
            self._table_p("Monto", 7.2, bold=True, color="#ffffff", align="RIGHT"),
            self._table_p("Estado", 7.2, bold=True, color="#ffffff", align="CENTER"),
        ]]
        for r in rows:
            issued_at = r.get("issued_at")
            date_str = issued_at.strftime("%Y-%m-%d") if hasattr(issued_at, "strftime") else str(issued_at or "")
            data.append([
                self._table_p(date_str, 7, align="CENTER"),
                self._table_p(r.get("apartment_code", ""), 7, align="CENTER"),
                self._table_p(r.get("owner_name", ""), 7),
                self._table_p(r.get("period", ""), 7, align="CENTER"),
                self._table_p(r.get("reason", ""), 7),
                self._table_p(self._money(r.get("amount")), 7, align="RIGHT"),
                self._table_p(r.get("status", ""), 7, align="CENTER"),
            ])
        if len(data) == 1:
            data.append([
                self._table_p("-", 7, align="CENTER"),
                self._table_p("-", 7, align="CENTER"),
                self._table_p("Sin multas registradas", 7),
                self._table_p("-", 7, align="CENTER"),
                self._table_p("-", 7),
                self._table_p(self._money(0), 7, align="RIGHT"),
                self._table_p("-", 7, align="CENTER"),
            ])
        table = Table(
            data,
            colWidths=[width * 0.11, width * 0.09, width * 0.19, width * 0.10, width * 0.29, width * 0.11, width * 0.11],
            repeatRows=1,
        )
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (1, -1), "CENTER"),
            ("ALIGN", (2, 0), (2, -1), "LEFT"),
            ("ALIGN", (3, 0), (3, -1), "CENTER"),
            ("ALIGN", (4, 0), (4, -1), "LEFT"),
            ("ALIGN", (5, 0), (5, -1), "RIGHT"),
            ("ALIGN", (6, 0), (6, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, _PDF_BORDER),
            ("WORDWRAP", (0, 0), (-1, -1), "CJK"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]))
        story.append(table)
        self._append_signature_grid(story, width=width, building=building, document_tag="REPORTE-MULTAS", signer_name="Administración", signer_role="Administrador del edificio")
        footer = self._footer_callback(building, width)
        doc.build(story, onFirstPage=footer, onLaterPages=footer)
        return output.getvalue()

    async def fines_excel(self, period: Optional[str] = None, start_date: Optional[date] = None, end_date: Optional[date] = None, status: Optional[str] = None, reason: Optional[str] = None, search: Optional[str] = None) -> bytes:
        rows = await self._fines_report_rows(period, start_date, end_date, status, reason, search)
        wb = Workbook()
        ws = wb.active
        ws.title = "Multas"
        headers = ["Fecha", "Departamento", "Propietario", "Período", "Motivo", "Monto", "Estado"]
        self._style_excel_header(ws, headers)
        for row_idx, row in enumerate(rows, 2):
            values = [str(row.get("issued_at") or ""), row.get("apartment_code", ""), row.get("owner_name", ""), row.get("period", ""), row.get("reason", ""), float(row.get("amount", 0)), row.get("status", "")]
            for col, value in enumerate(values, 1):
                ws.cell(row=row_idx, column=col).value = value
        for col in "ABCDEFG":
            ws.column_dimensions[col].width = 18
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def owners_pdf(self, start_date: Optional[date] = None, end_date: Optional[date] = None, status: Optional[str] = None) -> bytes:
        rows = await self._owners_report_rows(start_date, end_date, status)
        building = await get_default_building_config(getattr(self._payment_repo, "_conn", None))
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        story.extend(await self._pdf_header("Reporte de Propietarios", f"Rango: {self._date_label(None, start_date, end_date)} | Estado: {status or 'Todos'}", width=doc.width))
        story.extend([Paragraph(f"Total de propietarios: {len(rows)}", styles["Heading3"]), Spacer(1, 0.15*inch)])
        data = [["Ingreso", "Propietario", "Documento", "Email", "Teléfono", "Unidades", "Estado"]]
        data.extend([[str(r.get("created_at").date() if r.get("created_at") else ""), r.get("full_name", ""), r.get("document_id", ""), r.get("email", ""), r.get("phone", ""), r.get("units") or "", r.get("status", "")] for r in rows])
        table = Table(data, colWidths=[0.7*inch, 1.25*inch, 0.85*inch, 1.35*inch, 0.8*inch, 0.9*inch, 0.65*inch], repeatRows=1)
        table.setStyle(self._table_style(7))
        story.append(table)
        self._append_signature_grid(story, width=doc.width, building=building, document_tag="REPORTE-PROPIETARIOS", signer_name="Administración", signer_role="Administrador del edificio")
        story.extend([Spacer(1, 0.2 * inch), build_pdf_footer_bar(building, width=doc.width)])
        doc.build(story)
        return output.getvalue()

    async def owners_excel(self, start_date: Optional[date] = None, end_date: Optional[date] = None, status: Optional[str] = None) -> bytes:
        rows = await self._owners_report_rows(start_date, end_date, status)
        wb = Workbook()
        ws = wb.active
        ws.title = "Propietarios"
        headers = ["Ingreso", "Propietario", "Documento", "Email", "Teléfono", "Unidades", "Estado"]
        self._style_excel_header(ws, headers)
        for row_idx, row in enumerate(rows, 2):
            values = [str(row.get("created_at").date() if row.get("created_at") else ""), row.get("full_name", ""), row.get("document_id", ""), row.get("email", ""), row.get("phone", ""), row.get("units") or "", row.get("status", "")]
            for col, value in enumerate(values, 1):
                ws.cell(row=row_idx, column=col).value = value
        for col in "ABCDEFG":
            ws.column_dimensions[col].width = 20
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def buildings_pdf(self, start_date: Optional[date] = None, end_date: Optional[date] = None) -> bytes:
        rows = await self._buildings_report_rows(start_date, end_date)
        building = await get_default_building_config(getattr(self._payment_repo, "_conn", None))
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        story.extend(await self._pdf_header("Reporte de Edificios", f"Rango: {self._date_label(None, start_date, end_date)}", width=doc.width))
        story.extend([Paragraph(f"Total de edificios: {len(rows)}", styles["Heading3"]), Spacer(1, 0.15*inch)])
        data = [["Creado", "Edificio", "Dirección", "Teléfono", "Email", "Departamentos"]]
        data.extend([[str(r.get("created_at").date() if r.get("created_at") else ""), r.get("name", ""), r.get("address", ""), r.get("phone", ""), r.get("email", ""), str(r.get("apartments_count", 0))] for r in rows])
        table = Table(data, colWidths=[0.75*inch, 1.35*inch, 1.5*inch, 0.85*inch, 1.4*inch, 0.75*inch], repeatRows=1)
        table.setStyle(self._table_style(7))
        story.append(table)
        self._append_signature_grid(story, width=doc.width, building=building, document_tag="REPORTE-EDIFICIOS", signer_name="Administración", signer_role="Administrador del edificio")
        story.extend([Spacer(1, 0.2 * inch), build_pdf_footer_bar(building, width=doc.width)])
        doc.build(story)
        return output.getvalue()

    async def buildings_excel(self, start_date: Optional[date] = None, end_date: Optional[date] = None) -> bytes:
        rows = await self._buildings_report_rows(start_date, end_date)
        wb = Workbook()
        ws = wb.active
        ws.title = "Edificios"
        headers = ["Creado", "Edificio", "Dirección", "Teléfono", "Email", "Departamentos"]
        self._style_excel_header(ws, headers)
        for row_idx, row in enumerate(rows, 2):
            values = [str(row.get("created_at").date() if row.get("created_at") else ""), row.get("name", ""), row.get("address", ""), row.get("phone", ""), row.get("email", ""), int(row.get("apartments_count", 0))]
            for col, value in enumerate(values, 1):
                ws.cell(row=row_idx, column=col).value = value
        for col in "ABCDEF":
            ws.column_dimensions[col].width = 22
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def delinquency_excel(self) -> bytes:
        owners = await self._delinquency.list_owners()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Morosidad"
        
        # Headers
        headers = ["Propietario", "Email", "Documento", "Deuda Total", "Períodos Vencidos", "Estado"]
        header_fill = PatternFill(start_color="123C7A", end_color="123C7A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row, o in enumerate(owners, 2):
            ws.cell(row=row, column=1).value = o["owner_name"]
            ws.cell(row=row, column=2).value = o.get("email") or ""
            ws.cell(row=row, column=3).value = o["document_id"]
            ws.cell(row=row, column=4).value = float(o["deuda_total"])
            ws.cell(row=row, column=5).value = o["periodos_vencidos"]
            ws.cell(row=row, column=6).value = self._delinquency_status_label(o["status"])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 12
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def owner_ficha_pdf(self, owner_id: UUID) -> bytes:
        # 1. Fetch data
        conn = self._payment_repo._conn
        owner = await conn.fetchrow("SELECT * FROM owners WHERE id = $1", owner_id)
        if not owner:
            raise ValueError("Propietario no encontrado")
        owner = dict(owner)
        
        building = await get_default_building_config(conn)
        
        # Get first apartment
        apt = await conn.fetchrow(
            """
            SELECT a.* FROM apartments a
            JOIN owner_apartments oa ON a.id = oa.apartment_id
            WHERE oa.owner_id = $1
            ORDER BY oa.is_primary DESC, a.code
            LIMIT 1
            """,
            owner_id,
        )
        apt = dict(apt) if apt else {}

        # 2. Financial calculation
        total_pagado = await conn.fetchval(
            "SELECT COALESCE(SUM(amount), 0.0) FROM payments WHERE owner_id = $1 AND status = 'REGISTRADO'",
            owner_id,
        )
        last_payment_date = await conn.fetchval(
            "SELECT paid_at FROM payments WHERE owner_id = $1 AND status = 'REGISTRADO' ORDER BY paid_at DESC LIMIT 1",
            owner_id,
        )
        
        # Calculate balance
        balance_row = await conn.fetchrow(
            """
            SELECT
                COALESCE(SUM(fees_amount - payments_amount + fines_amount), 0.0) as total_balance
            FROM (
                SELECT
                    COALESCE(af.amount, 0.0) as fees_amount,
                    COALESCE(p.amount, 0.0) as payments_amount,
                    COALESCE(f.amount, 0.0) as fines_amount
                FROM apartment_fees af
                FULL OUTER JOIN payments p ON p.apartment_id = af.apartment_id AND p.period = af.period AND p.status = 'REGISTRADO' AND p.fine_id IS NULL
                FULL OUTER JOIN fines f ON f.apartment_id = af.apartment_id AND f.period = af.period AND f.status = 'ACTIVA'
                JOIN owner_apartments oa ON af.apartment_id = oa.apartment_id
                WHERE oa.owner_id = $1
            ) balance_calc
            """,
            owner_id,
        )
        saldo_actual = float(balance_row["total_balance"]) if balance_row else 0.0
        
        # Current month charges and payments
        current_month = datetime.now().strftime("%Y-%m")
        pagos_mes = await conn.fetchval(
            "SELECT COALESCE(SUM(amount), 0.0) FROM payments WHERE owner_id = $1 AND TO_CHAR(paid_at, 'YYYY-MM') = $2 AND status = 'REGISTRADO'",
            owner_id,
            current_month,
        )
        cargos_mes = await conn.fetchval(
            """
            SELECT COALESCE(SUM(af.amount), 0.0) FROM apartment_fees af
            JOIN owner_apartments oa ON af.apartment_id = oa.apartment_id
            WHERE oa.owner_id = $1 AND af.period = $2
            """,
            owner_id,
            current_month,
        )
        saldo_anterior = saldo_actual - float(cargos_mes) + float(pagos_mes)
        
        # Recent 3 payments
        recent_payments = await conn.fetch(
            """
            SELECT paid_at, period, amount
            FROM payments
            WHERE owner_id = $1 AND status = 'REGISTRADO'
            ORDER BY paid_at DESC
            LIMIT 3
            """,
            owner_id,
        )
        recent_payments = [dict(p) for p in recent_payments]
        
        # Settings
        due_day = 5
        settings_row = await conn.fetchrow("SELECT due_day FROM settings LIMIT 1")
        if settings_row:
            due_day = settings_row["due_day"]
        
        # Next due date
        now = datetime.now()
        if now.month == 12:
            next_due = date(now.year + 1, 1, due_day)
        else:
            next_due = date(now.year, now.month + 1, due_day)

        # 3. Setup PDF document
        output = io.BytesIO()
        width = A4[0] - 1.4 * cm
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            leftMargin=0.7 * cm,
            rightMargin=0.7 * cm,
            topMargin=0.5 * cm,
            bottomMargin=1.4 * cm,
        )
        story = []
        
        # Emission date and sheet number block
        emission_date = datetime.now().strftime("%d/%m/%Y")
        sheet_number = f"FTN-{datetime.now().year}-{str(owner['document_id'])[-6:].zfill(6)}"
        story.extend(
            build_pdf_brand_header(
                "FICHA DEL COPROPIETARIO",
                f"Ficha N.°: {sheet_number}",
                building,
                width=width,
            )
        )

        header_info = Table(
            [
                [
                    Paragraph(
                        "<font size='8' color='#123c7a'><b>DATOS DE EMISION</b></font><br/>"
                        f"<font size='8' color='#4b5563'>Fecha de emisión: {emission_date}</font><br/>"
                        f"<font size='8' color='#4b5563'><b>Ficha N.°: {sheet_number}</b></font>",
                        ParagraphStyle("FichaHeaderInfo", fontName="Helvetica", leading=11),
                    ),
                ]
            ],
            colWidths=[width],
        )
        header_info.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#123c7a")),
            ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d4dfef")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(header_info)
        story.append(Spacer(1, 0.08 * cm))
        
        # Helper for sections titles
        def make_section_title(title_text: str):
            t = Table(
                [[Paragraph(f"<font size='9' color='#123c7a'><b>{title_text}</b></font>", ParagraphStyle("SecTitle", fontName="Helvetica-Bold", leading=11))]],
                colWidths=[width]
            )
            t.setStyle(TableStyle([
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#123c7a")),
            ]))
            return t

        # SECTION 1: DATOS DEL COPROPIETARIO
        story.append(make_section_title("1. DATOS DEL COPROPIETARIO"))
        story.append(Spacer(1, 0.15 * cm))
        
        # Left side: owner info
        avatar_placeholder = Paragraph("<font size='28' color='#9ca3af'><b>👤</b></font>", ParagraphStyle("Avatar", fontName="Helvetica", alignment=1))
        
        reg_date_str = self._spanish_date(owner.get("created_at").date()) if owner.get("created_at") else "--"
        birth_date_str = self._spanish_date(owner.get("birth_date")) if owner.get("birth_date") else "--"
        
        owner_text = (
            f"<font size='10' color='#1f2937'><b>{escape(owner['full_name'])}</b></font><br/>"
            f"<font size='7.5' color='#6b7280'>Copropietario</font><br/><br/>"
            f"<font size='7.5' color='#4b5563'>✉ {escape(owner.get('email') or '--')}</font><br/>"
            f"<font size='7.5' color='#4b5563'>📞 {escape(owner.get('phone') or '--')}</font><br/>"
            f"<font size='7.5' color='#4b5563'>📇 C.I.: {escape(owner.get('document_id') or '--')}</font><br/>"
            f"<font size='7.5' color='#4b5563'>📅 Fecha de registro: {reg_date_str}</font>"
        )
        owner_info_p = Paragraph(owner_text, ParagraphStyle("OwnerInfoText", fontName="Helvetica", leading=11))
        
        # Right side: unit general info
        apt_code = apt.get("code") or "--"
        apt_tower = apt.get("tower") or "--"
        apt_floor = str(apt.get("floor") or "--")
        apt_area = f"{apt.get('area_sqm'):,.2f} m²" if apt.get("area_sqm") else "--"
        apt_quota = f"{apt.get('allocated_quota_percent'):,.2f} %" if apt.get("allocated_quota_percent") else "--"
        apt_status_label = "Al día" if saldo_actual <= 0 else "Con deuda"
        apt_status_color = "#1f8f4d" if saldo_actual <= 0 else "#c74444"
        
        unit_summary_data = [
            [Paragraph("<font size='7.5' color='#4b5563'>🏢 Departamento:</font>", ParagraphStyle("L1")), Paragraph(f"<font size='7.5' color='#1f2937'><b>{apt_code}</b></font>", ParagraphStyle("R1"))],
            [Paragraph("<font size='7.5' color='#4b5563'>🗼 Torre:</font>", ParagraphStyle("L2")), Paragraph(f"<font size='7.5' color='#1f2937'>{apt_tower}</font>", ParagraphStyle("R2"))],
            [Paragraph("<font size='7.5' color='#4b5563'>📍 Piso:</font>", ParagraphStyle("L3")), Paragraph(f"<font size='7.5' color='#1f2937'>{apt_floor}</font>", ParagraphStyle("R3"))],
            [Paragraph("<font size='7.5' color='#4b5563'>📐 Área del depto:</font>", ParagraphStyle("L4")), Paragraph(f"<font size='7.5' color='#1f2937'>{apt_area}</font>", ParagraphStyle("R4"))],
            [Paragraph("<font size='7.5' color='#4b5563'>% Porcentaje alícuota:</font>", ParagraphStyle("L5")), Paragraph(f"<font size='7.5' color='#1f2937'>{apt_quota}</font>", ParagraphStyle("R5"))],
            [Paragraph("<font size='7.5' color='#4b5563'>📊 Estado:</font>", ParagraphStyle("L6")), Paragraph(f"<font size='7.5' color='{apt_status_color}'><b>● {apt_status_label}</b></font>", ParagraphStyle("R6"))],
        ]
        
        unit_summary_table = Table(unit_summary_data, colWidths=[width * 0.22, width * 0.22])
        unit_summary_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("LINEBELOW", (0, 0), (-1, -2), 0.4, colors.HexColor("#e5e7eb")),
        ]))
        
        # Combine Left & Right
        sec1_table = Table(
            [[avatar_placeholder, owner_info_p, unit_summary_table]],
            colWidths=[1.8 * cm, width * 0.48, width * 0.44]
        )
        sec1_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ("BACKGROUND", (2, 0), (2, 0), colors.HexColor("#f8fafc")),
            ("BOX", (2, 0), (2, 0), 0.5, colors.HexColor("#e5e7eb")),
        ]))
        story.append(sec1_table)
        story.append(Spacer(1, 0.12 * cm))
        
        # SECTION 2: INFORMACIÓN DEL DEPARTAMENTO
        story.append(make_section_title("2. INFORMACIÓN DEL DEPARTAMENTO"))
        story.append(Spacer(1, 0.15 * cm))
        
        # Replicating department details table (two-column key-value format without image)
        apt_bedrooms = str(apt.get("bedrooms") or "3")
        apt_bathrooms = f"{float(apt.get('bathrooms') or 2.5):g}"
        apt_parking = apt.get("parking") or "1 (P-28)"
        apt_storage = apt.get("storage") or "B-12"
        apt_acquisition = self._spanish_date(apt.get("acquisition_date")) if apt.get("acquisition_date") else "15 de junio de 2022"
        apt_use = apt.get("use_type") or "Departamento residencial"
        
        sec2_data = [
            [
                Paragraph("<font size='7.5' color='#4b5563'>Código del departamento:</font>", ParagraphStyle("K1")),
                Paragraph(f"<font size='7.5' color='#1f2937'><b>{apt_code}</b></font>", ParagraphStyle("V1")),
                Paragraph("<font size='7.5' color='#4b5563'>Bodega:</font>", ParagraphStyle("K4")),
                Paragraph(f"<font size='7.5' color='#1f2937'>{apt_storage}</font>", ParagraphStyle("V4"))
            ],
            [
                Paragraph("<font size='7.5' color='#4b5563'>Tipo de unidad:</font>", ParagraphStyle("K2")),
                Paragraph(f"<font size='7.5' color='#1f2937'>{apt_use}</font>", ParagraphStyle("V2")),
                Paragraph("<font size='7.5' color='#4b5563'>Fecha de adquisición:</font>", ParagraphStyle("K5")),
                Paragraph(f"<font size='7.5' color='#1f2937'>{apt_acquisition}</font>", ParagraphStyle("V5"))
            ],
            [
                Paragraph("<font size='7.5' color='#4b5563'>Número de habitaciones:</font>", ParagraphStyle("K3")),
                Paragraph(f"<font size='7.5' color='#1f2937'>{apt_bedrooms}</font>", ParagraphStyle("V3")),
                Paragraph("<font size='7.5' color='#4b5563'>Uso del departamento:</font>", ParagraphStyle("K6")),
                Paragraph(f"<font size='7.5' color='#1f2937'>{apt_use}</font>", ParagraphStyle("V6"))
            ],
            [
                Paragraph("<font size='7.5' color='#4b5563'>Número de baños:</font>", ParagraphStyle("K3_1")),
                Paragraph(f"<font size='7.5' color='#1f2937'>{apt_bathrooms}</font>", ParagraphStyle("V3_1")),
                Paragraph("<font size='7.5' color='#4b5563'>Estado actual:</font>", ParagraphStyle("K7")),
                Paragraph(f"<font size='7.5' color='{apt_status_color}'><b>● {apt_status_label}</b></font>", ParagraphStyle("V7"))
            ],
            [
                Paragraph("<font size='7.5' color='#4b5563'>Parqueaderos:</font>", ParagraphStyle("K3_2")),
                Paragraph(f"<font size='7.5' color='#1f2937'>{apt_parking}</font>", ParagraphStyle("V3_2")),
                Spacer(1,1), Spacer(1,1)
            ]
        ]
        sec2_table = Table(sec2_data, colWidths=[width * 0.28, width * 0.22, width * 0.28, width * 0.22])
        sec2_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("LEFTPADDING", (0, 0), (-1, -1), 1),
            ("RIGHTPADDING", (0, 0), (-1, -1), 1),
            ("LINEBELOW", (0, 0), (-1, -1), 0.2, colors.HexColor("#e5e7eb")),
        ]))
        story.append(sec2_table)
        story.append(Spacer(1, 0.12 * cm))
        
        # SECTION 3: INFORMACIÓN FINANCIERA
        story.append(make_section_title("3. INFORMACIÓN FINANCIERA"))
        story.append(Spacer(1, 0.15 * cm))
        
        # Column 1: Resumen de pagos
        last_pay_str = self._spanish_date(last_payment_date) if last_payment_date else "--"
        next_due_str = self._spanish_date(next_due)
        
        resumen_data = [
            [Paragraph("<font size='8' color='#123c7a'><b>Resumen de pagos</b></font>", ParagraphStyle("ResT"))],
            [Paragraph(f"<font size='7' color='#4b5563'>Estado actual:</font><br/><font size='8.5' color='{apt_status_color}'><b>{apt_status_label}</b></font>", ParagraphStyle("Item1"))],
            [Paragraph(f"<font size='7' color='#4b5563'>Último pago:</font><br/><font size='7.5' color='#1f2937'>{last_pay_str}</font>", ParagraphStyle("Item2"))],
            [Paragraph(f"<font size='7' color='#4b5563'>Próximo vencimiento:</font><br/><font size='7.5' color='#1f2937'>{next_due_str}</font>", ParagraphStyle("Item3"))],
            [Paragraph(f"<font size='7' color='#4b5563'>Total alícuotas pagadas:</font><br/><font size='8' color='#1f2937'><b>{self._money(total_pagado)}</b></font>", ParagraphStyle("Item4"))],
            [Paragraph(f"<font size='7' color='#4b5563'>Total en mora:</font><br/><font size='8.5' color='{apt_status_color}'><b>{self._money(saldo_actual)}</b></font>", ParagraphStyle("Item5"))],
            [Paragraph(f"<font size='6.5' color='#1f8f4d'><b>✓ El copropietario se encuentra al día con sus obligaciones.</b></font>" if saldo_actual <= 0 else f"<font size='6.5' color='#c74444'><b>⚠️ El copropietario registra saldo pendiente de pago.</b></font>", ParagraphStyle("AlertText"))]
        ]
        resumen_table = Table(resumen_data, colWidths=[width * 0.32])
        resumen_table.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("LINEBELOW", (0, 0), (0, 0), 0.6, colors.HexColor("#123c7a")),
            ("BACKGROUND", (0, -1), (0, -1), colors.HexColor("#edf9f0") if saldo_actual <= 0 else colors.HexColor("#fdf2f2")),
            ("BOX", (0, -1), (0, -1), 0.3, colors.HexColor("#1f8f4d") if saldo_actual <= 0 else colors.HexColor("#f87171")),
        ]))
        
        # Column 2: Últimos 3 pagos registrados
        pagos_headers = [
            Paragraph("<b>Fecha</b>", ParagraphStyle("PH1", fontName="Helvetica-Bold", size=7, color="#ffffff")),
            Paragraph("<b>Concepto</b>", ParagraphStyle("PH2", fontName="Helvetica-Bold", size=7, color="#ffffff")),
            Paragraph("<b>Valor (USD)</b>", ParagraphStyle("PH3", fontName="Helvetica-Bold", size=7, color="#ffffff", align="RIGHT")),
        ]
        pagos_rows = [pagos_headers]
        
        sum_last_3 = Decimal("0.00")
        for p in recent_payments:
            pay_date = p["paid_at"].strftime("%d/%m/%Y") if isinstance(p["paid_at"], (date, datetime)) else str(p["paid_at"])
            pay_period = self._period_name(p["period"]).capitalize()
            pay_concept = f"Alícuota - {pay_period}"
            pay_amount = Decimal(str(p["amount"]))
            sum_last_3 += pay_amount
            
            pagos_rows.append([
                Paragraph(f"<font size='7' color='#374151'>{pay_date}</font>", ParagraphStyle("PD1")),
                Paragraph(f"<font size='7' color='#374151'>{pay_concept}</font>", ParagraphStyle("PD2")),
                Paragraph(f"<font size='7' color='#374151'>{self._money(pay_amount)}</font>", ParagraphStyle("PD3", align="RIGHT")),
            ])
            
        while len(pagos_rows) < 4:
            pagos_rows.append([
                Paragraph("<font size='7' color='#9ca3af'>--</font>", ParagraphStyle("PD1")),
                Paragraph("<font size='7' color='#9ca3af'>Sin registro</font>", ParagraphStyle("PD2")),
                Paragraph("<font size='7' color='#9ca3af'>$0.00</font>", ParagraphStyle("PD3", align="RIGHT")),
            ])
            
        # Total row
        pagos_rows.append([
            Paragraph("<b>Total pagado en el período</b>", ParagraphStyle("PTT", fontName="Helvetica-Bold", size=7, color="#123c7a")),
            Spacer(1,1),
            Paragraph(f"<b>{self._usd(sum_last_3)}</b>", ParagraphStyle("PTA", fontName="Helvetica-Bold", size=7, color="#123c7a", align="RIGHT")),
        ])
        
        pagos_table = Table(pagos_rows, colWidths=[width * 0.095, width * 0.135, width * 0.10])
        pagos_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123c7a")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -2), 0.3, colors.HexColor("#e5e7eb")),
            ("SPAN", (0, -1), (1, -1)),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#dbe7f7")),
            ("BOX", (0, -1), (-1, -1), 0.5, colors.HexColor("#123c7a")),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ]))
        
        # Column 3: Estado de cuenta actual
        fin_rows = [
            [Paragraph("<b>Concepto</b>", ParagraphStyle("FH1", fontName="Helvetica-Bold", size=7, color="#ffffff")), Paragraph("<b>Monto (USD)</b>", ParagraphStyle("FH2", fontName="Helvetica-Bold", size=7, color="#ffffff", align="RIGHT"))],
            [Paragraph("<font size='7' color='#4b5563'>Saldo anterior:</font>", ParagraphStyle("FL1")), Paragraph(f"<font size='7' color='#1f2937'>{self._money(saldo_anterior)}</font>", ParagraphStyle("FV1", align="RIGHT"))],
            [Paragraph("<font size='7' color='#4b5563'>Cargos del mes:</font>", ParagraphStyle("FL2")), Paragraph(f"<font size='7' color='#1f2937'>{self._money(cargos_mes)}</font>", ParagraphStyle("FV2", align="RIGHT"))],
            [Paragraph("<font size='7' color='#4b5563'>Pagos del mes:</font>", ParagraphStyle("FL3")), Paragraph(f"<font size='7' color='#c74444'>-{self._money(pagos_mes)}</font>", ParagraphStyle("FV3", align="RIGHT"))],
            [Paragraph("<b>Saldo actual:</b>", ParagraphStyle("FL4", fontName="Helvetica-Bold", size=7.5, color="#123c7a")), Paragraph(f"<b>{self._money(saldo_actual)}</b>", ParagraphStyle("FV4", fontName="Helvetica-Bold", size=7.5, color=apt_status_color, align="RIGHT"))]
        ]
        fin_table = Table(fin_rows, colWidths=[width * 0.18, width * 0.15])
        fin_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123c7a")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -2), 0.3, colors.HexColor("#e5e7eb")),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#edf9f0") if saldo_actual <= 0 else colors.HexColor("#fdf2f2")),
            ("BOX", (0, -1), (-1, -1), 0.5, colors.HexColor("#1f8f4d") if saldo_actual <= 0 else colors.HexColor("#f87171")),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ]))
        
        # Combine the 3 columns into Section 3 table
        sec3_table = Table([[resumen_table, pagos_table, fin_table]], colWidths=[width * 0.34, width * 0.33, width * 0.33])
        sec3_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(sec3_table)
        story.append(Spacer(1, 0.10 * cm))
        
        # SECTION 4: CONTACTOS DE EMERGENCIA
        story.append(make_section_title("4. CONTACTOS DE EMERGENCIA"))
        story.append(Spacer(1, 0.15 * cm))
        
        # Three cards
        contact1_text = (
            f"<font size='8' color='#123c7a'><b>Contacto principal</b></font><br/><br/>"
            f"<font size='7.5' color='#1f2937'><b>{escape(owner['full_name'])}</b></font><br/>"
            f"<font size='7' color='#4b5563'>📞 {escape(owner.get('phone') or '--')}</font><br/>"
            f"<font size='7' color='#4b5563'>✉ {escape(owner.get('email') or '--')}</font>"
        )
        contact2_text = (
            f"<font size='8' color='#123c7a'><b>Contacto alterno</b></font><br/><br/>"
            f"<font size='7.5' color='#1f2937'><b>{escape(owner.get('occupant_name') or '--')}</b></font><br/>"
            f"<font size='7' color='#4b5563'>Relación: {escape(owner.get('occupant_relation') or '--')}</font><br/>"
            f"<font size='7' color='#4b5563'>📞 {escape(owner.get('occupant_phone') or '--')}</font>"
        )
        contact3_text = (
            f"<font size='8' color='#123c7a'><b>Uso exclusivo de emergencia</b></font><br/><br/>"
            f"<font size='7.5' color='#1f2937'><b>Administración</b></font><br/>"
            f"<font size='7' color='#4b5563'>📞 +593 2 398 4500</font><br/>"
            f"<font size='7' color='#4b5563'>✉ administracion@torresnetanya.com</font>"
        )
        
        c1_table = Table([[Paragraph(contact1_text, ParagraphStyle("C1T", leading=10))]], colWidths=[width * 0.32])
        c1_table.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#e5e7eb")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ]))
        
        c2_table = Table([[Paragraph(contact2_text, ParagraphStyle("C2T", leading=10))]], colWidths=[width * 0.32])
        c2_table.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#e5e7eb")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ]))
        
        c3_table = Table([[Paragraph(contact3_text, ParagraphStyle("C3T", leading=10))]], colWidths=[width * 0.32])
        c3_table.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#e5e7eb")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ]))
        
        sec4_table = Table([[c1_table, c2_table, c3_table]], colWidths=[width * 0.34, width * 0.33, width * 0.33])
        sec4_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(sec4_table)
        story.append(Spacer(1, 0.08 * cm))
        
        # SECTION 5: OBSERVACIONES
        story.append(make_section_title("5. OBSERVACIONES"))
        story.append(Spacer(1, 0.15 * cm))
        
        obs_box = Table(
            [[Paragraph("<font size='7.5' color='#4b5563'>Ninguna observación registrada.</font>", ParagraphStyle("ObsText"))]],
            colWidths=[width]
        )
        obs_box.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#e5e7eb")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(obs_box)
        story.append(Spacer(1, 0.14 * cm))
        
        self._append_signature_grid(
            story,
            width=width,
            building=building,
            document_tag=f"FICHA-{sheet_number}",
            signer_name=owner.get("full_name") or "Copropietario",
            signer_role="Copropietario",
        )
        story.append(Spacer(1, 0.18 * cm))
        
        def draw_footer(canvas, doc):
            canvas.saveState()
            footer = build_pdf_footer_bar(building, width=doc.width, page_text=f"Página {doc.page}")
            _, footer_height = footer.wrap(doc.width, doc.bottomMargin)
            footer.drawOn(canvas, doc.leftMargin, doc.bottomMargin - footer_height)
            canvas.restoreState()

        doc.build(story, onFirstPage=draw_footer, onLaterPages=draw_footer)
        return output.getvalue()
